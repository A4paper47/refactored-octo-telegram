import os
import json
import asyncio
import logging
import threading
import sys
import re
import platform
import tempfile
import zipfile
from datetime import datetime, timezone, timedelta
from io import BytesIO, StringIO
import csv

from openpyxl import Workbook

from jinja2 import TemplateNotFound

from flask import Flask, request, redirect, url_for, render_template, flash, Response, jsonify, send_file, session
from flask_login import LoginManager, login_user, login_required, logout_user, UserMixin
from werkzeug.security import generate_password_hash, check_password_hash
from sqlalchemy import func, text as sql_text, or_, inspect
from telegram import Update

from db import init_db, db
from models import (
    AdminUser,
    AdminTelegramUser,
    Assignment,
    TranslationSubmission,
    VOTeam,
    Movie,
    Translator,
    VORoleSubmission,
    TranslationTask,
    AppKV,
    GroupOpenRequest,
    MovieEvent,
    MovieAlias,
)
from models import GroupMovieContext, GroupRoleImportRequest
from export_excel import export_simple
from export_dynamic import export_excel_dynamic, backup_json_zip_dynamic
from restore_dynamic import restore_dry_run as restore_dry_run_dynamic, restore_from_backup_zip
from bot_ptb import (
    build_bot,
    build_admin_digest_text,
    _normalize_priority_mode as normalize_priority_mode,
    _priority_mode_hours as priority_mode_hours,
    _priority_mode_urgent_only as priority_mode_urgent_only,
    _priority_mode_deadline as priority_mode_deadline,
    _priority_mode_label as priority_mode_label,
    _load_import_req_roles,
    _load_import_req_suggestions,
    _refresh_role_import_request,
    _auto_assign_movie_roles,
    upsert_movie,
    find_repairable_movie_titles,
    repair_movie_title_db,
    find_movie_aliases,
    add_movie_alias_db,
    delete_movie_alias_db,
    parse_movie_from_filename,
    _parse_movie_from_role_helper_filename,
    _clean_movie_title_candidate,
    _resolve_movie_query,
)
from ops_log import log_event, fetch_logs
from movie_history import record_movie_event, fetch_movie_history, fetch_recent_movie_events
from movie_merge import duplicate_groups, merge_preview, merge_movies, merge_simulation
from assign_logic import (
    parse_lines,
    role_gender,
    pick_vo,
    movie_load,
)

# --------------------------------------------------
# CONFIG
# --------------------------------------------------
BOT_TOKEN = os.getenv("BOT_TOKEN")
BACKUP_TELEGRAM_CHAT_ID = (os.getenv("BACKUP_TELEGRAM_CHAT_ID") or "").strip()
CRON_SECRET = (os.getenv("CRON_SECRET") or "").strip()
WEBHOOK_SECRET = os.getenv("WEBHOOK_SECRET", "secret")
SECRET_KEY = os.getenv("SECRET_KEY", "devkey")

# Temporary: disable web login page (for emergency access / debugging)
DISABLE_LOGIN = os.getenv("DISABLE_LOGIN", "1").strip() not in ("0", "false", "False", "")

# Optional access key when login is disabled (recommended for security).
# If set, web UI requires unlocking once per browser session.
ADMIN_PANEL_KEY = (os.getenv("ADMIN_PANEL_KEY") or "").strip()

# --------------------------------------------------
# SIMPLE APP KV (DB-based settings, no extra services)
# --------------------------------------------------
def kv_get(key: str) -> str:
    """Read a small config value from DB (app_kv table). Returns '' if missing."""
    try:
        row = AppKV.query.filter_by(key=key).first()
        return (row.value or "").strip() if row else ""
    except Exception:
        try:
            db.session.rollback()
        except Exception:
            pass
        return ""


def kv_set(key: str, value: str) -> bool:
    """Upsert a small config value into DB (app_kv table)."""
    try:
        row = AppKV.query.filter_by(key=key).first()
        if not row:
            row = AppKV(key=key, value=value)
        else:
            row.value = value
        db.session.add(row)
        db.session.commit()
        return True
    except Exception:
        try:
            db.session.rollback()
        except Exception:
            pass
        return False


def get_backup_chat_id(override: str = "") -> str:
    """Backup destination chat id: override > ENV > DB."""
    return (override or "").strip() or BACKUP_TELEGRAM_CHAT_ID or kv_get("backup_chat_id")


MOVIE_CODE_RE = re.compile(r"^[A-Za-z]{2,5}-\d{6}-\d{2}$")


def _slug_lang(lang: str | None) -> str:
    v = (lang or "").strip().lower()
    if not v:
        return "bn"
    mapping = {
        "bn": "bn",
        "bangla": "bn",
        "bengali": "bn",
        "bengal": "bn",
        "en": "en",
        "eng": "en",
        "english": "en",
        "ms": "ms",
        "bm": "ms",
        "malay": "ms",
        "melayu": "ms",
        "bahasa melayu": "ms",
    }
    return mapping.get(v, v[:8])


def _looks_like_movie_code(value: str | None) -> bool:
    return bool(MOVIE_CODE_RE.fullmatch((value or "").strip()))


def _make_movie_code(lang: str | None) -> str:
    lang2 = _slug_lang(lang).upper()
    day = datetime.utcnow().strftime("%y%m%d")
    prefix = f"{lang2}-{day}-"
    n = Movie.query.filter(Movie.code.like(f"{prefix}%")).count() + 1
    return f"{prefix}{n:02d}"


def _movie_display_title(title: str | None, year: str | int | None) -> str:
    t = (title or "").strip()
    y = str(year).strip() if year is not None else ""
    return f"{t} ({y})" if t and y and y.lower() != "none" else (t or y or "")


def _active_movie_expr():
    return or_(Movie.is_archived.is_(False), Movie.is_archived.is_(None))


def _archived_movie_expr():
    return Movie.is_archived.is_(True)


def _archive_movie_record(movie: Movie, clear_active_rows: bool = True) -> None:
    movie.is_archived = True
    movie.archived_at = datetime.utcnow()
    movie.status = "ARCHIVED"
    movie.updated_at = datetime.utcnow()
    movie.translator_assigned = None
    if clear_active_rows:
        Assignment.query.filter((Assignment.movie_id == movie.id) | (Assignment.project == movie.code)).delete(synchronize_session=False)
        VORoleSubmission.query.filter_by(movie=movie.code).delete(synchronize_session=False)
        TranslationTask.query.filter((TranslationTask.movie_id == movie.id) | (TranslationTask.movie_code == movie.code)).delete(synchronize_session=False)


def _hard_delete_movie_record(movie: Movie) -> None:
    code = (movie.code or "").strip()
    mid = movie.id
    Assignment.query.filter((Assignment.movie_id == mid) | (Assignment.project == code)).delete(synchronize_session=False)
    VORoleSubmission.query.filter_by(movie=code).delete(synchronize_session=False)
    TranslationTask.query.filter((TranslationTask.movie_id == mid) | (TranslationTask.movie_code == code)).delete(synchronize_session=False)
    TranslationSubmission.query.filter((TranslationSubmission.movie_id == mid) | (TranslationSubmission.movie == code)).delete(synchronize_session=False)
    GroupOpenRequest.query.filter((GroupOpenRequest.movie_id == mid) | (GroupOpenRequest.movie_code == code)).delete(synchronize_session=False)
    db.session.delete(movie)


def _find_movie_by_title(title: str, year: str | None = None, lang: str | None = None):
    q = Movie.query.filter(_active_movie_expr()).filter(func.lower(Movie.title) == (title or "").strip().lower())
    if year:
        q = q.filter(Movie.year == str(year).strip())
    if lang:
        q = q.filter(Movie.lang == _slug_lang(lang))
    movie = q.order_by(Movie.id.desc()).first()
    if movie:
        return movie

    # Loose fallback when year/lang were not provided or older rows were incomplete.
    q2 = Movie.query.filter(_active_movie_expr()).filter(Movie.title.ilike((title or "").strip()))
    if year:
        q2 = q2.filter(Movie.year == str(year).strip())
    return q2.order_by(Movie.id.desc()).first()



def _find_movie_any(query: str):
    raw = (query or "").strip()
    if not raw:
        return None
    if _looks_like_movie_code(raw):
        return Movie.query.filter_by(code=raw.upper()).first()
    normalized = raw.lower()
    movie = Movie.query.filter(func.lower(Movie.title) == normalized).order_by(Movie.id.desc()).first()
    if movie:
        return movie
    return Movie.query.filter(Movie.title.ilike(f"%{raw}%")).order_by(Movie.id.desc()).first()

def _resolve_or_create_assignment_movie(project_raw: str, title_raw: str, year_raw: str | None, lang_raw: str | None):
    """
    Accept either:
    - project_raw = movie code
    - project_raw = movie title
    - title_raw = movie title

    Returns: (movie, created_movie, title_first_mode)
    """
    project_raw = (project_raw or "").strip()
    title_raw = (title_raw or "").strip()
    year = (str(year_raw).strip() if year_raw not in (None, "") else None) or None
    lang = _slug_lang(lang_raw or "bn")

    # Legacy code-first flow still supported.
    if _looks_like_movie_code(project_raw):
        code = project_raw.upper()
        movie = Movie.query.filter_by(code=code).first()
        if movie:
            changed = False
            if getattr(movie, "is_archived", False):
                movie.is_archived = False
                movie.archived_at = None
                if (movie.status or "").upper() == "ARCHIVED":
                    movie.status = "RECEIVED"
                changed = True
            if title_raw and title_raw != movie.title:
                movie.title = title_raw
                changed = True
            if year and not (movie.year or "").strip():
                movie.year = year
                changed = True
            if lang and not (movie.lang or "").strip():
                movie.lang = lang
                changed = True
            if changed:
                movie.updated_at = datetime.utcnow()
                db.session.commit()
            return movie, False, False

        movie = Movie(
            code=code,
            title=(title_raw or code),
            year=year,
            lang=lang,
            status="VO_ASSIGNED",
        )
        db.session.add(movie)
        db.session.commit()
        return movie, True, False

    # Title-first flow: if project box contains title, use it.
    title = title_raw or project_raw
    if not title:
        raise ValueError("Movie name or code required")

    movie = _find_movie_by_title(title, year=year, lang=lang)
    if movie:
        if getattr(movie, "is_archived", False):
            movie.is_archived = False
            movie.archived_at = None
            if (movie.status or "").upper() == "ARCHIVED":
                movie.status = "RECEIVED"
            movie.updated_at = datetime.utcnow()
            db.session.commit()
        return movie, False, True

    movie = Movie(
        code=_make_movie_code(lang),
        title=title,
        year=year,
        lang=lang,
        status="VO_ASSIGNED",
    )
    db.session.add(movie)
    db.session.commit()
    return movie, True, True


ADMIN_EMAIL = os.getenv("ADMIN_EMAIL", "admin@local")
ADMIN_PASSWORD = os.getenv("ADMIN_PASSWORD", "admin123")

DROP_CHAT_ID = os.getenv("DROP_CHAT_ID")
ADMIN_TELEGRAM_CHAT_ID = os.getenv("ADMIN_TELEGRAM_CHAT_ID")

# Version is shared between web and bot.
from version import APP_VERSION, BUILD_ID
RENDER_EXTERNAL_URL = os.getenv("RENDER_EXTERNAL_URL", "")

BOT_ENABLED = bool(BOT_TOKEN)
if not BOT_ENABLED:
    logging.warning("BOT_TOKEN missing — Telegram bot disabled; web UI will still run.")

logging.basicConfig(level=logging.INFO)

# Security: redact bot tokens from logs + silence noisy HTTP client logs
from sec_logging import install_security_logging
install_security_logging()
log = logging.getLogger("app")

# Malaysia Time (MYT) is UTC+8.
MYT_OFFSET_HOURS = 8


def utc_to_myt(dt: datetime | None) -> datetime | None:
    if not dt:
        return None
    return dt + timedelta(hours=MYT_OFFSET_HOURS)


def fmt_myt(dt: datetime | None) -> str:
    if not dt:
        return "-"
    return utc_to_myt(dt).strftime("%Y-%m-%d %H:%M") + " MYT"


def parse_myt_datetime_local(val: str | None) -> datetime | None:
    """Parse HTML <input type=datetime-local> value as MYT and convert to UTC."""
    s = (val or "").strip()
    if not s:
        return None
    try:
        # 'YYYY-MM-DDTHH:MM' or 'YYYY-MM-DD HH:MM'
        s = s.replace(" ", "T")
        dt_local = datetime.strptime(s[:16], "%Y-%m-%dT%H:%M")
        return dt_local - timedelta(hours=MYT_OFFSET_HOURS)
    except Exception:
        return None

# --------------------------------------------------
# APP INIT
# --------------------------------------------------
app = Flask(__name__)
app.secret_key = SECRET_KEY
app.url_map.strict_slashes = False
# Flask-Login: bypass @login_required when DISABLE_LOGIN=true
app.config["LOGIN_DISABLED"] = bool(DISABLE_LOGIN)

init_db(app)

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = "login"


@app.context_processor
def _inject_flags():
    return {
        "login_disabled": bool(DISABLE_LOGIN),
        "panel_key_enabled": bool(ADMIN_PANEL_KEY),
        "backup_chat_configured": bool(BACKUP_TELEGRAM_CHAT_ID),
        "app_version": APP_VERSION,
        "build_id": BUILD_ID,
}

# --------------------------------------------------
# TELEGRAM: one persistent asyncio loop + initialized PTB application
# --------------------------------------------------
bot_app = None
_bot_loop = None


def run_bot_coro(coro):
    raise RuntimeError("Telegram bot disabled (missing BOT_TOKEN).")


# ✅ compatibility alias (your code previously called run_async)
def run_async(coro):
    return run_bot_coro(coro)


if BOT_ENABLED:
    bot_app = build_bot(BOT_TOKEN)
    _bot_loop = asyncio.new_event_loop()

    def _bot_thread_target():
        asyncio.set_event_loop(_bot_loop)
        _bot_loop.run_forever()

    threading.Thread(target=_bot_thread_target, daemon=True).start()

    async def _bot_startup():
        # PTB21 requires initialize() before process_update
        await bot_app.initialize()
        await bot_app.start()

    asyncio.run_coroutine_threadsafe(_bot_startup(), _bot_loop).result(timeout=30)

    def run_bot_coro(coro):
        fut = asyncio.run_coroutine_threadsafe(coro, _bot_loop)
        return fut.result(timeout=30)

    def run_async(coro):
        return run_bot_coro(coro)

# --------------------------------------------------
# DB: migrate helpers + seed
# --------------------------------------------------
DEFAULT_VO_TEAM = [
    # name, gender, level, speed, urgent_ok, active
    ("Rafiq", "male", "expert_old", "normal", True, True),
    ("Ahamad", "male", "expert_old", "normal", True, True),
    ("Tirthankar", "male", "expert_old", "normal", True, True),
    ("Sanu", "male", "expert_old", "normal", True, True),

    ("Rezual", "male", "trained_new", "normal", True, True),
    ("Shihab", "male", "trained_new", "normal", True, True),
    ("Himu", "male", "trained_new", "normal", True, True),
    ("Sohel", "male", "trained_new", "normal", True, True),
    ("Morshed", "male", "trained_new", "normal", True, True),
    ("Ahsan", "male", "trained_new", "normal", True, True),
    ("Shawon", "male", "trained_new", "normal", True, True),
    ("Rabbi", "male", "new_limited", "normal", False, True),

    ("Shazia", "female", "expert_old", "normal", True, True),
    ("Sraboni", "female", "expert_old", "normal", True, True),
    ("Sharmim", "female", "trained_new", "normal", True, True),
    ("Kashkeya", "female", "trained_new", "normal", True, True),
    ("Joyashree", "female", "trained_new", "normal", True, True),
    ("Labanya", "female", "trained_new", "slow", True, True),
    ("Rafiq-f", "female", "new_limited", "normal", False, True),
]

DEFAULT_TRANSLATORS = [
    "Ryan",
    "Shafaytul",
    "Suman",
    "Samael",
    "Ezaz",
    "Adeeb",
    "Lamia",
    "Ananna",
    "Sumi",
    "Monira",
]


def _table_exists(table: str) -> bool:
    """Dialect-safe table existence check.

    SQLAlchemy inspector can be flaky across some managed Postgres setups
    (schema/search_path quirks). We use Postgres-native checks when possible.
    """
    try:
        dialect = db.engine.dialect.name
        if dialect != "sqlite":
            # to_regclass respects search_path (usually includes public)
            val = db.session.execute(sql_text("SELECT to_regclass(:t)"), {"t": table}).scalar()
            return val is not None
        insp = inspect(db.engine)
        return table in insp.get_table_names()
    except Exception:
        return False


def _col_exists(table: str, col: str) -> bool:
    """Dialect-safe column existence check (Postgres + sqlite)."""
    try:
        dialect = db.engine.dialect.name
        if dialect != "sqlite":
            row = db.session.execute(
                sql_text(
                    """
                    SELECT 1
                    FROM information_schema.columns
                    WHERE table_name = :t
                      AND column_name = :c
                      AND table_schema = ANY (current_schemas(false))
                    LIMIT 1
                    """
                ),
                {"t": table, "c": col},
            ).first()
            return row is not None

        # sqlite pragma
        if not table.replace("_", "").isalnum():
            return False
        res = db.session.execute(sql_text(f"PRAGMA table_info({table})")).all()
        return any(r[1] == col for r in res)
    except Exception:
        return False


def _safe_exec(ddl: str, params: dict | None = None) -> bool:
    """Execute SQL safely (never abort startup migrations)."""
    try:
        db.session.execute(sql_text(ddl), params or {})
        db.session.commit()
        return True
    except Exception as e:
        db.session.rollback()
        log.warning("MIGRATE failed: %s | %s", ddl, e)
        return False


def _add_column(table: str, col: str, col_type_sql: str, default_sql: str | None = None):
    """Add a column with best-effort IF NOT EXISTS on Postgres."""
    dialect = db.engine.dialect.name
    if dialect != "sqlite":
        ddl = f"ALTER TABLE {table} ADD COLUMN IF NOT EXISTS {col} {col_type_sql}"
        if default_sql:
            ddl += f" DEFAULT {default_sql}"
        _safe_exec(ddl)
        return

    # sqlite: no IF NOT EXISTS
    if not _col_exists(table, col):
        ddl = f"ALTER TABLE {table} ADD COLUMN {col} {col_type_sql}"
        if default_sql:
            ddl += f" DEFAULT {default_sql}"
        _safe_exec(ddl)


def _ensure_group_role_import_table():
    """Best-effort creation for the role import queue table.

    On some managed deployments, relying only on db.create_all() can leave this
    page broken if a worker boots against an older DB schema. We create the table
    and backfill missing columns defensively so /role_imports never hard-crashes.
    """
    dialect = db.engine.dialect.name
    if dialect == "sqlite":
        _safe_exec(
            """
            CREATE TABLE IF NOT EXISTS group_role_import_request (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              tg_chat_id BIGINT,
              tg_message_id BIGINT,
              title TEXT,
              year TEXT,
              lang TEXT,
              roles_text TEXT,
              roles_json TEXT,
              suggested_json TEXT,
              status TEXT,
              created_at DATETIME,
              expires_at DATETIME,
              requested_by_tg_id BIGINT,
              requested_by_name TEXT,
              reviewed_by_tg_id BIGINT,
              reviewed_by_name TEXT,
              reviewed_at DATETIME,
              note TEXT
            )
            """
        )
    else:
        _safe_exec(
            """
            CREATE TABLE IF NOT EXISTS group_role_import_request (
              id BIGSERIAL PRIMARY KEY,
              tg_chat_id BIGINT,
              tg_message_id BIGINT,
              title VARCHAR(255),
              year VARCHAR(10),
              lang VARCHAR(30),
              roles_text TEXT,
              roles_json TEXT,
              suggested_json TEXT,
              status VARCHAR(20),
              created_at TIMESTAMPTZ DEFAULT CURRENT_TIMESTAMP,
              expires_at TIMESTAMPTZ,
              requested_by_tg_id BIGINT,
              requested_by_name VARCHAR(120),
              reviewed_by_tg_id BIGINT,
              reviewed_by_name VARCHAR(120),
              reviewed_at TIMESTAMPTZ,
              note TEXT
            )
            """
        )
        _safe_exec("CREATE INDEX IF NOT EXISTS ix_group_role_import_request_status ON group_role_import_request (status)")
        _safe_exec("CREATE INDEX IF NOT EXISTS ix_group_role_import_request_created_at ON group_role_import_request (created_at)")
        _safe_exec("CREATE INDEX IF NOT EXISTS ix_group_role_import_request_expires_at ON group_role_import_request (expires_at)")
        _safe_exec("CREATE INDEX IF NOT EXISTS ix_group_role_import_request_tg_chat_id ON group_role_import_request (tg_chat_id)")

    # Column backfill for partially-created tables.
    _add_column("group_role_import_request", "tg_chat_id", "BIGINT")
    _add_column("group_role_import_request", "tg_message_id", "BIGINT")
    _add_column("group_role_import_request", "title", "VARCHAR(255)")
    _add_column("group_role_import_request", "year", "VARCHAR(10)")
    _add_column("group_role_import_request", "lang", "VARCHAR(30)")
    _add_column("group_role_import_request", "roles_text", "TEXT")
    _add_column("group_role_import_request", "roles_json", "TEXT")
    _add_column("group_role_import_request", "suggested_json", "TEXT")
    _add_column("group_role_import_request", "status", "VARCHAR(20)")
    _add_column("group_role_import_request", "created_at", "TIMESTAMP")
    _add_column("group_role_import_request", "expires_at", "TIMESTAMP")
    _add_column("group_role_import_request", "requested_by_tg_id", "BIGINT")
    _add_column("group_role_import_request", "requested_by_name", "VARCHAR(120)")
    _add_column("group_role_import_request", "reviewed_by_tg_id", "BIGINT")
    _add_column("group_role_import_request", "reviewed_by_name", "VARCHAR(120)")
    _add_column("group_role_import_request", "reviewed_at", "TIMESTAMP")
    _add_column("group_role_import_request", "note", "TEXT")


def _role_import_feature_ready() -> bool:
    return _table_exists("group_role_import_request")


def auto_migrate_schema():
    """Add new columns safely without dropping tables.

    Render/Postgres can boot multiple workers; we guard migrations with a
    Postgres advisory lock to avoid race conditions.
    """

    dialect = db.engine.dialect.name
    lock_key = 823_771_991
    got_lock = False

    if dialect != "sqlite":
        try:
            db.session.execute(sql_text("SELECT pg_advisory_lock(:k)"), {"k": lock_key})
            db.session.commit()
            got_lock = True
        except Exception:
            db.session.rollback()

    try:
        # --- small schema additions ---
        _add_column("assignment", "movie_id", "INTEGER")
        _add_column("assignment", "urgent", "BOOLEAN", default_sql="TRUE")
        _add_column("assignment", "priority_mode", "VARCHAR(20)")
        _add_column("assignment", "deadline_at", "TIMESTAMP")
        _add_column("assignment", "last_reminded_at", "TIMESTAMP")
        _add_column("assignment", "remind_count", "INTEGER", default_sql="0")
        _add_column("translation_submission", "movie_id", "INTEGER")
        _add_column("translation_submission", "telegram_event_id", "BIGINT")
        _add_column("movie", "is_archived", "BOOLEAN", default_sql="FALSE")
        _add_column("movie", "archived_at", "TIMESTAMP")
        _add_column("translation_submission", "tg_chat_id", "BIGINT")
        _add_column("translation_submission", "tg_message_id", "BIGINT")

        # --- pending role import queue ---
        _ensure_group_role_import_table()

        # --- movie_event: lightweight per-movie action timeline ---
        if dialect == "sqlite":
            _safe_exec(
                """
                CREATE TABLE IF NOT EXISTS movie_event (
                  id INTEGER PRIMARY KEY AUTOINCREMENT,
                  movie_id INTEGER,
                  movie_code TEXT,
                  movie_title TEXT,
                  event_type TEXT,
                  summary TEXT NOT NULL,
                  detail TEXT,
                  actor_source TEXT,
                  actor_name TEXT,
                  created_at DATETIME
                )
                """
            )
        else:
            _safe_exec(
                """
                CREATE TABLE IF NOT EXISTS movie_event (
                  id BIGSERIAL PRIMARY KEY,
                  movie_id INTEGER,
                  movie_code VARCHAR(40),
                  movie_title VARCHAR(255),
                  event_type VARCHAR(40),
                  summary TEXT NOT NULL,
                  detail TEXT,
                  actor_source VARCHAR(40),
                  actor_name VARCHAR(120),
                  created_at TIMESTAMPTZ DEFAULT CURRENT_TIMESTAMP
                )
                """
            )
        if _table_exists("movie_event"):
            _add_column("movie_event", "movie_id", "INTEGER")
            _add_column("movie_event", "movie_code", "VARCHAR(40)")
            _add_column("movie_event", "movie_title", "VARCHAR(255)")
            _add_column("movie_event", "event_type", "VARCHAR(40)")
            _add_column("movie_event", "summary", "TEXT")
            _add_column("movie_event", "detail", "TEXT")
            _add_column("movie_event", "actor_source", "VARCHAR(40)")
            _add_column("movie_event", "actor_name", "VARCHAR(120)")
            _add_column("movie_event", "created_at", "TIMESTAMP")

        # --- app_kv: tiny key/value config store (e.g. backup destination chat id) ---
        if dialect == "sqlite":
            _safe_exec(
                """
                CREATE TABLE IF NOT EXISTS app_kv (
                  key TEXT PRIMARY KEY,
                  value TEXT,
                  updated_at DATETIME
                )
                """
            )
        else:
            _safe_exec(
                """
                CREATE TABLE IF NOT EXISTS app_kv (
                  key VARCHAR(80) PRIMARY KEY,
                  value TEXT,
                  updated_at TIMESTAMPTZ DEFAULT CURRENT_TIMESTAMP
                )
                """
            )
        if _table_exists("app_kv"):
            _add_column("app_kv", "value", "TEXT")
            _add_column("app_kv", "updated_at", "TIMESTAMP")


        # --- translator roster (legacy safety) ---
        if _table_exists("translator"):
            _add_column("translator", "name", "VARCHAR(120)")
            _add_column("translator", "tg_user_id", "BIGINT")
            _add_column("translator", "tg_username", "VARCHAR(120)")
            _add_column("translator", "active", "BOOLEAN", default_sql="TRUE")
            _add_column("translator", "languages", "VARCHAR(80)")
            _add_column("translator", "note", "TEXT")
            _add_column("translator", "last_seen_at", "TIMESTAMP")
            _add_column("translator", "created_at", "TIMESTAMP")
            _add_column("translator", "updated_at", "TIMESTAMP")

            # Legacy: tg_user_id NOT NULL
            if _col_exists("translator", "tg_user_id") and dialect != "sqlite":
                _safe_exec("ALTER TABLE translator ALTER COLUMN tg_user_id DROP NOT NULL")

            # Backfill missing names
            if _col_exists("translator", "name"):
                if _col_exists("translator", "tg_username"):
                    _safe_exec(
                        """
                        UPDATE translator
                        SET name = COALESCE(NULLIF(name,''), NULLIF(tg_username,''), 'Translator-' || id)
                        WHERE name IS NULL OR name = ''
                        """
                    )
                else:
                    _safe_exec(
                        """
                        UPDATE translator
                        SET name = COALESCE(NULLIF(name,''), 'Translator-' || id)
                        WHERE name IS NULL OR name = ''
                        """
                    )

        # --- movie table columns ---
        if _table_exists("movie"):
            _add_column("movie", "movie_code", "VARCHAR(40)")
            _add_column("movie", "code", "VARCHAR(40)")
            _add_column("movie", "translator_assigned", "VARCHAR(120)")
            _add_column("movie", "movie_card_chat_id", "BIGINT")
            _add_column("movie", "movie_card_message_id", "BIGINT")
            _add_column("movie", "received_at", "TIMESTAMP")
            _add_column("movie", "submitted_at", "TIMESTAMP")
            _add_column("movie", "completed_at", "TIMESTAMP")
            _add_column("movie", "vo_group_chat_id", "BIGINT")
            _add_column("movie", "vo_group_invite_link", "TEXT")
            _add_column("movie", "created_at", "TIMESTAMP")
            _add_column("movie", "updated_at", "TIMESTAMP")

            # Best-effort code backfill
            try:
                has_code = _col_exists("movie", "code")
                has_movie_code = _col_exists("movie", "movie_code")
                if has_movie_code:
                    db.session.execute(
                        sql_text(
                            """
                            UPDATE movie
                            SET movie_code = COALESCE(NULLIF(code,''), 'MOV-' || id)
                            WHERE (movie_code IS NULL OR movie_code = '')
                            """
                        )
                    )
                if has_code:
                    db.session.execute(
                        sql_text(
                            """
                            UPDATE movie
                            SET code = COALESCE(NULLIF(movie_code,''), 'MOV-' || id)
                            WHERE (code IS NULL OR code = '')
                            """
                        )
                    )
                if _col_exists("movie", "created_at"):
                    db.session.execute(
                        sql_text(
                            """
                            UPDATE movie
                            SET created_at = COALESCE(created_at, updated_at, received_at, CURRENT_TIMESTAMP)
                            WHERE created_at IS NULL
                            """
                        )
                    )
                db.session.commit()
            except Exception:
                db.session.rollback()

            if _col_exists("movie", "code"):
                _safe_exec("CREATE INDEX IF NOT EXISTS ix_movie_code ON movie(code)")
            if _col_exists("movie", "movie_code"):
                _safe_exec("CREATE INDEX IF NOT EXISTS ix_movie_movie_code ON movie(movie_code)")

        # --- vo_role_submission safety ---
        if _table_exists("vo_role_submission"):
            _add_column("vo_role_submission", "movie", "VARCHAR(200)")
            _add_column("vo_role_submission", "vo", "VARCHAR(120)")
            _add_column("vo_role_submission", "role", "VARCHAR(20)")
            _add_column("vo_role_submission", "lines", "INTEGER", default_sql="0")
            _add_column("vo_role_submission", "submitted_at", "TIMESTAMP")
            _add_column("vo_role_submission", "tg_chat_id", "BIGINT")
            _add_column("vo_role_submission", "tg_message_id", "BIGINT")
            _add_column("vo_role_submission", "media_type", "VARCHAR(20)")
            _add_column("vo_role_submission", "file_id", "VARCHAR(200)")
            _add_column("vo_role_submission", "file_name", "VARCHAR(255)")

            # Backfill from legacy column(s)
            if _col_exists("vo_role_submission", "movie") and _col_exists("vo_role_submission", "project"):
                _safe_exec(
                    """
                    UPDATE vo_role_submission
                    SET movie = COALESCE(NULLIF(movie,''), project)
                    WHERE movie IS NULL OR movie = ''
                    """
                )

        # --- vo_team optional telegram columns (for reminders) ---
        if _table_exists("vo_team"):
            _add_column("vo_team", "tg_user_id", "BIGINT")
            _add_column("vo_team", "tg_username", "VARCHAR(120)")
            _add_column("vo_team", "last_seen_at", "TIMESTAMP")

        # --- translation_task table (translator workload) ---
        # IMPORTANT: some older deployments created translation_task WITHOUT movie_id.
        # If the ORM model includes movie_id, any SELECT will fail until the column exists.
        # So we (1) ensure table exists, then (2) add missing columns best-effort.
        if dialect == "sqlite":
            _safe_exec(
                """
                CREATE TABLE IF NOT EXISTS translation_task (
                  id INTEGER PRIMARY KEY AUTOINCREMENT,
                  movie_id INTEGER,
                  movie_code TEXT,
                  title TEXT,
                  year TEXT,
                  lang TEXT,
                  translator_id INTEGER,
                  translator_name TEXT,
                  status TEXT NOT NULL DEFAULT 'SENT',
                  deadline_at DATETIME,
                  sent_at DATETIME,
                  completed_at DATETIME,
                  last_reminded_at DATETIME,
                  remind_count INTEGER NOT NULL DEFAULT 0,
                  created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                  updated_at DATETIME
                )
                """
            )
        else:
            _safe_exec(
                """
                CREATE TABLE IF NOT EXISTS translation_task (
                  id SERIAL PRIMARY KEY,
                  movie_id INTEGER,
                  movie_code VARCHAR(40),
                  title TEXT,
                  year VARCHAR(10),
                  lang VARCHAR(30),
                  translator_id INTEGER,
                  translator_name VARCHAR(120),
                  status VARCHAR(20) NOT NULL DEFAULT 'SENT',
                  deadline_at TIMESTAMPTZ,
                  sent_at TIMESTAMPTZ,
                  completed_at TIMESTAMPTZ,
                  last_reminded_at TIMESTAMPTZ,
                  remind_count INTEGER NOT NULL DEFAULT 0,
                  created_at TIMESTAMPTZ DEFAULT CURRENT_TIMESTAMP,
                  updated_at TIMESTAMPTZ
                )
                """
            )

        # Column-level safety for partially-created translation_task tables
        if _table_exists("translation_task"):
            _add_column("translation_task", "movie_id", "INTEGER")
            _add_column("translation_task", "movie_code", "VARCHAR(40)")
            _add_column("translation_task", "title", "TEXT")
            _add_column("translation_task", "year", "VARCHAR(10)")
            _add_column("translation_task", "lang", "VARCHAR(30)")
            _add_column("translation_task", "translator_id", "INTEGER")
            _add_column("translation_task", "translator_name", "VARCHAR(120)")
            _add_column("translation_task", "status", "VARCHAR(20)", default_sql="'SENT'")
            _add_column("translation_task", "priority_mode", "VARCHAR(20)")
            _add_column("translation_task", "deadline_at", "TIMESTAMP")
            _add_column("translation_task", "sent_at", "TIMESTAMP")
            _add_column("translation_task", "completed_at", "TIMESTAMP")
            _add_column("translation_task", "last_reminded_at", "TIMESTAMP")
            _add_column("translation_task", "remind_count", "INTEGER", default_sql="0")
            _add_column("translation_task", "created_at", "TIMESTAMP")
            _add_column("translation_task", "updated_at", "TIMESTAMP")

        # Best-effort indexes
        _safe_exec("CREATE INDEX IF NOT EXISTS ix_translation_task_translator_id ON translation_task(translator_id)")
        _safe_exec("CREATE INDEX IF NOT EXISTS ix_translation_task_status ON translation_task(status)")
        _safe_exec("CREATE INDEX IF NOT EXISTS ix_translation_task_deadline ON translation_task(deadline_at)")

        # --- system logs table (for web pop-out console) ---
        if dialect == "sqlite":
            _safe_exec(
                """
                CREATE TABLE IF NOT EXISTS system_logs (
                  id INTEGER PRIMARY KEY AUTOINCREMENT,
                  ts DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
                  level TEXT NOT NULL,
                  source TEXT,
                  message TEXT NOT NULL,
                  traceback TEXT
                )
                """
            )
        else:
            _safe_exec(
                """
                CREATE TABLE IF NOT EXISTS system_logs (
                  id SERIAL PRIMARY KEY,
                  ts TIMESTAMPTZ NOT NULL DEFAULT CURRENT_TIMESTAMP,
                  level VARCHAR(12) NOT NULL,
                  source VARCHAR(64),
                  message TEXT NOT NULL,
                  traceback TEXT
                )
                """
            )

    finally:
        if got_lock and dialect != "sqlite":
            _safe_exec("SELECT pg_advisory_unlock(:k)", {"k": lock_key})


def seed_vo_team():
    # Best-effort rename for legacy spelling (keeps your old roster consistent)
    rename_map = {"Rezaul": "Rezual", "Sharmin": "Sharmim"}
    for old, new in rename_map.items():
        try:
            old_row = VOTeam.query.filter_by(name=old).first()
            if old_row and not VOTeam.query.filter_by(name=new).first():
                old_row.name = new
        except Exception:
            pass
    try:
        db.session.commit()
    except Exception:
        db.session.rollback()

    added = 0
    for name, gender, level, speed, urgent_ok, active in DEFAULT_VO_TEAM:
        row = VOTeam.query.filter_by(name=name).first()
        if row:
            row.gender = gender
            row.level = level
            row.speed = speed
            row.urgent_ok = urgent_ok
            row.active = active
        else:
            db.session.add(
                VOTeam(
                    name=name,
                    gender=gender,
                    level=level,
                    speed=speed,
                    urgent_ok=urgent_ok,
                    active=active,
                )
            )
            added += 1
    db.session.commit()
    if added:
        log.info("VO seed complete. added=%s", added)


def seed_translators():
    added = 0
    # Avoid query-triggered autoflush while seeding (safer on partially-migrated DBs).
    with db.session.no_autoflush:
        for name in DEFAULT_TRANSLATORS:
            row = Translator.query.filter_by(name=name).first()
            if row:
                # keep existing info (username, tg id, etc)
                if row.active is None:
                    row.active = True
            else:
                db.session.add(Translator(name=name, active=True))
                added += 1
    db.session.commit()
    if added:
        log.info("Translator seed complete. added=%s", added)


def backfill_translation_tasks(limit: int = 500):
    """Best-effort backfill so existing deployments get TranslationTask rows.

    Rules:
      - Only for movies with translator_assigned.
      - Skip if a task already exists for (movie_id) or (movie_code).
      - If we can match a Translator roster row, attach translator_id.
    """
    try:
        # Only run if table exists
        if not _table_exists("translation_task"):
            return

        q = Movie.query.filter(Movie.translator_assigned.isnot(None)).order_by(Movie.id.desc()).limit(limit)
        movies = [m for m in q.all() if (m.translator_assigned or "").strip()]
        if not movies:
            return

        created = 0
        for m in movies:
            # task exists?
            exists = None
            if m.id:
                exists = TranslationTask.query.filter_by(movie_id=m.id).first()
            if not exists and (m.code or "").strip():
                exists = TranslationTask.query.filter_by(movie_code=m.code).first()
            if exists:
                continue

            who = (m.translator_assigned or "").strip()
            who_norm = who.lstrip("@").strip()
            tr = None
            if who_norm:
                tr = Translator.query.filter(Translator.tg_username.ilike(who_norm)).first()
            if not tr and who_norm:
                tr = Translator.query.filter(Translator.name.ilike(who_norm)).first()
            if not tr and who:
                # fuzzy: any translator name/username contained in string
                for t in Translator.query.all():
                    if t.tg_username and t.tg_username.lower() in who.lower():
                        tr = t
                        break
                    if t.name and t.name.lower() in who.lower():
                        tr = t
                        break

            sent_at = m.received_at or m.created_at or m.updated_at
            db.session.add(
                TranslationTask(
                    movie_id=m.id,
                    movie_code=m.code,
                    title=m.title,
                    year=m.year,
                    lang=m.lang,
                    translator_id=tr.id if tr else None,
                    translator_name=tr.name if tr else who,
                    status="SENT",
                    sent_at=sent_at,
                )
            )
            created += 1

        if created:
            db.session.commit()
            log.info("Backfilled translation_task rows=%s", created)
    except Exception as e:
        db.session.rollback()
        log.warning("Backfill translation_task failed: %s", e)


with app.app_context():
    db.create_all()
    auto_migrate_schema()

    admin = AdminUser.query.filter_by(email=ADMIN_EMAIL).first()
    if not admin:
        db.session.add(AdminUser(email=ADMIN_EMAIL, password_hash=generate_password_hash(ADMIN_PASSWORD)))
        db.session.commit()
        log.info("Admin user created")
    else:
        # keep env password in sync (optional but useful)
        admin.password_hash = generate_password_hash(ADMIN_PASSWORD)
        db.session.commit()

    seed_vo_team()
    seed_translators()
    backfill_translation_tasks()

# --------------------------------------------------
# AUTH
# --------------------------------------------------
class AdminLogin(UserMixin):
    def __init__(self, row: AdminUser):
        self.id = str(row.id)
        self.email = row.email


@login_manager.user_loader
def load_user(user_id):
    row = db.session.get(AdminUser, int(user_id))
    return AdminLogin(row) if row else None


def require_admin(fn):
    def wrapper(*args, **kwargs):
        return fn(*args, **kwargs)
    wrapper.__name__ = fn.__name__
    return wrapper


@app.get("/logs")
@login_required
def logs_feed():
    limit = int(request.args.get("limit", "60"))
    return jsonify({"logs": fetch_logs(limit=limit)})


@app.post("/logs/clear")
@login_required
def logs_clear():
    """Clear system logs from the web UI."""
    mode = (request.form.get("mode") or "all").strip().lower()
    if mode == "errors":
        # Delete error/warn style logs (best-effort, case-insensitive).
        db.session.execute(
            sql_text(
                "DELETE FROM system_logs "
                "WHERE lower(level) LIKE '%error%' "
                "   OR lower(level) LIKE '%warn%' "
                "   OR lower(level) LIKE '%tg.error%'"
            )
        )
    else:
        db.session.execute(sql_text("DELETE FROM system_logs"))
    db.session.commit()
    return jsonify({"ok": True})


@app.post("/ops/run")
@login_required
def ops_run():
    """Run selected ops from the web UI and stream results to system logs.

    This is owner/admin gated by web login.
    """
    action = (request.form.get("action") or "").strip()

    if not BOT_ENABLED or not bot_app:
        flash("❌ Bot disabled (missing BOT_TOKEN)")
        log_event("WARN", "web.ops", "BOT_DISABLED")
        return redirect(request.referrer or url_for("dashboard"))

    ref = (request.form.get("ref") or "").strip()

    log_event("INFO", "web.ops", f"RUN action={action} ref={ref}")

    try:
        # Call bot-side ops helpers (coroutines) via the bot loop.
        from bot_ptb import (
            ops_request_group,
            ops_approve_request,
            ops_reject_request,
            ops_post_assignments,
        )

        if action == "request_group":
            run_async(ops_request_group(bot_app.bot, ref, actor="web"))
        elif action == "approve_request":
            run_async(ops_approve_request(bot_app.bot, int(ref), actor="web"))
        elif action == "reject_request":
            # Format: "<request_id>|<note>"
            if "|" not in ref:
                raise ValueError("Reject requires: <request_id>|<note>")
            rid, note = ref.split("|", 1)
            run_async(ops_reject_request(bot_app.bot, int(rid.strip()), note.strip(), actor="web"))
        elif action == "post_assignments":
            run_async(ops_post_assignments(bot_app.bot, ref, actor="web"))
        else:
            raise ValueError("Unknown action")

        log_event("INFO", "web.ops", f"OK action={action} ref={ref}")
        flash("✅ Command executed. Check Logs.")
    except Exception as e:
        import traceback as tb

        log_event("ERROR", "web.ops", f"FAIL action={action} ref={ref}: {e}", tb.format_exc())
        flash(f"❌ Failed: {e}")

    return redirect(url_for("telegram_panel"))

# --------------------------------------------------
# ROUTES: auth
# --------------------------------------------------
@app.route("/", methods=["GET", "POST"])
@app.route("/login", methods=["GET", "POST"])
def login():
    if DISABLE_LOGIN:
        return redirect(url_for("dashboard"))

    if request.method == "POST":
        email = request.form.get("email", "").strip()
        password = request.form.get("password", "").strip()

        row = AdminUser.query.filter_by(email=email).first()
        if row and check_password_hash(row.password_hash, password):
            login_user(AdminLogin(row))
            return redirect(url_for("dashboard"))

        flash("❌ Invalid login")

    return render_template("login.html")


@app.route("/logout")
@login_required
def logout():
    logout_user()
    return redirect(url_for("login"))


# --------------------------------------------------
# OPTIONAL PANEL UNLOCK (when login is disabled)
# --------------------------------------------------
@app.route("/unlock", methods=["GET", "POST"])
def unlock():
    """Unlock admin panel when DISABLE_LOGIN=true and ADMIN_PANEL_KEY is set.

    This keeps the "no-login" emergency mode while still protecting the UI from public access.
    """
    if not (DISABLE_LOGIN and ADMIN_PANEL_KEY):
        return redirect(url_for("dashboard"))

    if request.method == "POST":
        key = (request.form.get("key") or "").strip()
        if key and key == ADMIN_PANEL_KEY:
            session["panel_unlocked"] = True
            session["panel_unlocked_at"] = datetime.utcnow().isoformat()
            nxt = (request.args.get("next") or "").strip()
            return redirect(nxt or url_for("dashboard"))
        flash("❌ Wrong key")

    return render_template("unlock.html")


@app.get("/lock")
def lock_panel():
    try:
        session.pop("panel_unlocked", None)
        session.pop("panel_unlocked_at", None)
    except Exception:
        pass
    return redirect(url_for("unlock"))


@app.before_request
def _panel_key_gate():
    """Protect the UI when login is disabled.

    If ADMIN_PANEL_KEY is set, require the user to unlock once per session.
    Telegram webhook endpoints and health/version are always allowed.
    """
    if not (DISABLE_LOGIN and ADMIN_PANEL_KEY):
        return None

    # Always allow static assets
    if request.endpoint == "static" or request.path.startswith("/static/"):
        return None

    # Allow webhook / health / version / unlock
    if request.path.startswith("/webhook/"):
        return None
    if request.path in ("/health", "/version", "/unlock", "/lock"):
        return None

    # Allow favicon
    if request.path == "/favicon.ico":
        return None

    if session.get("panel_unlocked") is True:
        return None

    # Redirect to unlock, keep next
    return redirect(url_for("unlock", next=request.path))


@app.route("/health")
def health():
    """Lightweight health check (no heavy queries)."""
    db_ok = True
    db_error = ""
    try:
        db.session.execute(sql_text("SELECT 1"))
    except Exception as e:
        db_ok = False
        db_error = str(e)[:200]
        try:
            db.session.rollback()
        except Exception:
            pass
    return jsonify({
        "ok": True,
        "app_version": APP_VERSION,
        "build_id": BUILD_ID,
        "server_time_utc": datetime.now(timezone.utc).isoformat(),
        "bot_enabled": bool(BOT_TOKEN),
        "db_ok": db_ok,
        "db_error": db_error,
    })

# ✅ upgraded /version (so you can confirm correct build)
@app.route("/version")
def version():
    routes = []
    for rule in sorted(app.url_map.iter_rules(), key=lambda r: r.rule):
        if rule.endpoint == "static":
            continue
        routes.append({
            "rule": str(rule.rule),
            "methods": sorted([m for m in rule.methods if m not in ("HEAD", "OPTIONS")]),
            "endpoint": rule.endpoint,
        })

    return jsonify({
        "app_version": APP_VERSION,
        "build_id": BUILD_ID,
        "python": sys.version.split()[0],
        "platform": platform.platform(),
        "server_time_utc": datetime.now(timezone.utc).isoformat(),
        "app_file": __file__,
        "routes_count": len(routes),
        "routes": routes,
    })

@app.before_request
def _reset_failed_tx():
    """Postgres: if a previous statement failed, the transaction is aborted.
    Roll back so the next query in this request can run.
    """
    try:
        db.session.rollback()
    except Exception:
        pass



# --------------------------------------------------
# DASHBOARD
# --------------------------------------------------
@app.route("/dashboard")
@login_required
@require_admin
def dashboard():
    # --- legacy counters ---
    movies_count = db.session.query(func.count(func.distinct(Assignment.project))).scalar() or 0
    assignments_count = Assignment.query.count()
    subs_pending = TranslationSubmission.query.filter_by(status="READY_FOR_QA").count()

    latest = TranslationSubmission.query.order_by(TranslationSubmission.submitted_at.desc()).limit(10).all()

    now_utc = datetime.utcnow()

    # --- VO per-person summary (total roles + pending roles + late roles) ---
    vo_rows = VOTeam.query.filter_by(active=True).order_by(VOTeam.gender.asc(), VOTeam.name.asc()).all()
    # preload submissions for faster diff
    # key: (movie, vo, role)
    try:
        subs = VORoleSubmission.query.with_entities(
            VORoleSubmission.movie, VORoleSubmission.vo, VORoleSubmission.role
        ).all()
    except Exception as e:
        # If migrations haven't run yet on a legacy DB, avoid crashing the UI.
        log.warning("dashboard: VORoleSubmission columns missing? using empty submitted set: %s", e)
        try:
            db.session.rollback()
        except Exception:
            pass
        subs = []
    done_triplet = {(m, v, r) for (m, v, r) in subs}
    done_pair = {(m, r) for (m, v, r) in subs}

    vo_stats = []
    for vo in vo_rows:
        assigned = Assignment.query.filter_by(vo=vo.name).all()
        total_roles = len(assigned)
        pending_roles = 0
        late_roles = 0
        for a in assigned:
            is_done = ((a.project, a.vo, a.role) in done_triplet) or ((a.project, a.role) in done_pair)
            if not is_done:
                pending_roles += 1
                if a.deadline_at and a.deadline_at < now_utc:
                    late_roles += 1
        vo_stats.append({
            "id": vo.id,
            "name": vo.name,
            "gender": vo.gender,
            "total": int(total_roles),
            "pending": int(pending_roles),
            "late": int(late_roles),
        })

    vo_with_job = sum(1 for r in vo_stats if r["total"] > 0)
    vo_no_job = sum(1 for r in vo_stats if r["total"] == 0)
    vo_busy = sum(1 for r in vo_stats if r["pending"] > 0)
    vo_idle = sum(1 for r in vo_stats if r["total"] > 0 and r["pending"] == 0)

    # --- Translator per-person summary (TranslationTask) ---
    translators = Translator.query.order_by(Translator.name.asc()).all()
    # preload tasks
    tasks = []
    try:
        tasks = TranslationTask.query.all()
    except Exception as e:
        log.warning("dashboard: TranslationTask query failed (empty): %s", e)
    try:
        db.session.rollback()
    except Exception:
        pass
    tr_stats = []
    for t in translators:
        # match tasks by translator_id; fallback match by name/username if legacy rows exist
        rel = [x for x in tasks if (x.translator_id == t.id)]
        if not rel and (t.name or t.tg_username):
            keyset = set([ (t.name or "").strip().lower(), (t.tg_username or "").strip().lower() ])
            rel = [x for x in tasks if (x.translator_name or "").strip().lower() in keyset]

        total = len(rel)
        pending = sum(1 for x in rel if (x.status or "").upper() != "COMPLETED")
        late = sum(1 for x in rel if (x.status or "").upper() != "COMPLETED" and x.deadline_at and x.deadline_at < now_utc)
        tr_stats.append({
            "id": t.id,
            "name": t.name,
            "tg_username": t.tg_username,
            "total": int(total),
            "pending": int(pending),
            "late": int(late),
            "active": bool(t.active),
        })

    tr_with_job = sum(1 for r in tr_stats if r["total"] > 0)
    tr_no_job = sum(1 for r in tr_stats if r["total"] == 0)
    tr_busy = sum(1 for r in tr_stats if r["pending"] > 0)
    tr_idle = sum(1 for r in tr_stats if r["total"] > 0 and r["pending"] == 0)

    # --- Daily + monthly summary (db-agnostic) ---
    from datetime import timedelta

    today = datetime.utcnow().date()
    days = [today - timedelta(days=i) for i in range(13, -1, -1)]

    # Pull recent submissions once
    t_from = datetime.utcnow() - timedelta(days=14)
    tr_subs = TranslationSubmission.query.filter(TranslationSubmission.submitted_at >= t_from).all()
    try:
        vo_subs = VORoleSubmission.query.filter(VORoleSubmission.submitted_at >= t_from).all()
    except Exception as e:
        log.warning("dashboard: vo_subs query failed (empty): %s", e)
        vo_subs = []

    daily = []
    for d in days:
        t_cnt = sum(1 for s in tr_subs if s.submitted_at and s.submitted_at.date() == d)
        v_cnt = sum(1 for s in vo_subs if s.submitted_at and s.submitted_at.date() == d)
        daily.append({"date": d.isoformat(), "translation": t_cnt, "vo": v_cnt})

    # Monthly (last 6 months)
    def ym(dt):
        return dt.year, dt.month

    # build month keys
    months = []
    y, m = today.year, today.month
    for i in range(5, -1, -1):
        # go back i months
        yy, mm = y, m - i
        while mm <= 0:
            yy -= 1
            mm += 12
        months.append((yy, mm))

    # Pull 6 months data
    m_from = datetime.utcnow() - timedelta(days=190)
    tr_subs_m = TranslationSubmission.query.filter(TranslationSubmission.submitted_at >= m_from).all()
    try:
        vo_subs_m = VORoleSubmission.query.filter(VORoleSubmission.submitted_at >= m_from).all()
    except Exception as e:
        log.warning("dashboard: vo_subs_m query failed (empty): %s", e)
        vo_subs_m = []

    monthly = []
    for yy, mm in months:
        t_cnt = sum(1 for s in tr_subs_m if s.submitted_at and ym(s.submitted_at) == (yy, mm))
        v_cnt = sum(1 for s in vo_subs_m if s.submitted_at and ym(s.submitted_at) == (yy, mm))
        monthly.append({"month": f"{yy:04d}-{mm:02d}", "translation": t_cnt, "vo": v_cnt})

    recent_activity = []
    try:
        for ev in fetch_recent_movie_events(limit=18, include_archived=True):
            recent_activity.append({
                "ts": ev.created_at.strftime("%Y-%m-%d %H:%M") if ev.created_at else "-",
                "movie_code": (ev.movie_code or "").strip() or "-",
                "movie_title": (ev.movie_title or "").strip() or (ev.movie_code or "-"),
                "event_type": ev.event_type or "INFO",
                "summary": ev.summary or "",
                "detail": ev.detail or "",
                "actor": " / ".join([x for x in [ev.actor_source, ev.actor_name] if x]) or "-",
            })
    except Exception as e:
        log.warning("dashboard: recent activity query failed (empty): %s", e)
        try:
            db.session.rollback()
        except Exception:
            pass

    return render_template(
        "dashboard.html",
        movies_count=movies_count,
        assignments_count=assignments_count,
        subs_pending=subs_pending,
        latest=latest,
        vo_stats=vo_stats,
        tr_stats=tr_stats,
        vo_summary={"with_job": vo_with_job, "no_job": vo_no_job, "busy": vo_busy, "idle": vo_idle},
        tr_summary={"with_job": tr_with_job, "no_job": tr_no_job, "busy": tr_busy, "idle": tr_idle},
        daily=daily,
        monthly=monthly,
        recent_activity=recent_activity,
    )

# --------------------------------------------------
# ASSIGNMENTS
# --------------------------------------------------
@app.route("/archived")
@login_required
@require_admin
def archived_panel():
    q = (request.args.get("q") or "").strip()
    try:
        limit = max(1, min(100, int(request.args.get("limit") or 30)))
    except Exception:
        limit = 30

    rows_q = Movie.query.filter(_archived_movie_expr())
    if q:
        code = q.strip().upper()
        rows_q = rows_q.filter(or_(Movie.code == code, Movie.title.ilike(f"%{q}%")))

    movies = rows_q.order_by(Movie.archived_at.desc().nullslast(), Movie.id.desc()).limit(limit).all()
    rows = []
    for m in movies:
        subtitle_bits = []
        if m.year:
            subtitle_bits.append(str(m.year))
        if m.translator_assigned:
            subtitle_bits.append(f"TR {m.translator_assigned}")
        rows.append({
            "display_name": _movie_display_title(m.title, m.year),
            "subtitle": " • ".join([x for x in subtitle_bits if x]),
            "code": m.code,
            "lang": (m.lang or "").upper() or "-",
            "status": m.status or "ARCHIVED",
            "archived_at": m.archived_at.strftime("%Y-%m-%d %H:%M") if m.archived_at else "-",
        })
    return render_template("archived.html", rows=rows, q=q, limit=limit)


@app.route('/title_repair')
@login_required
@require_admin
def title_repair_panel():
    q = (request.args.get('q') or '').strip()
    try:
        limit = max(1, min(int(request.args.get('limit') or 50), 200))
    except Exception:
        limit = 50
    rows = find_repairable_movie_titles(q=q, limit=limit, include_archived=True)
    return render_template('title_repair.html', rows=rows, q=q, limit=limit)


@app.post('/title_repair/run')
@login_required
@require_admin
def title_repair_run():
    code = (request.form.get('movie_code') or '').strip().upper()
    q = (request.form.get('q') or '').strip()
    try:
        limit = max(1, min(int(request.form.get('limit') or 50), 200))
    except Exception:
        limit = 50
    movie = Movie.query.filter_by(code=code).first() if code else None
    if not movie:
        flash('❌ Movie not found for title repair')
        return redirect(url_for('title_repair_panel', q=q, limit=limit))
    try:
        result = repair_movie_title_db(movie, actor_source='web', actor_name='title_repair_run')
        if result.get('changed'):
            flash(f"✅ Title repaired for {movie.code}: {result.get('old_title')} → {result.get('new_title')}")
        elif result.get('reason') == 'conflict':
            issue = result.get('issue') or {}
            conflict = issue.get('conflict')
            flash(f"⚠️ Repair blocked for {movie.code}: conflict with {(conflict.code if conflict else 'existing movie')}")
        else:
            flash(f"ℹ️ No title repair needed for {movie.code}")
    except Exception as e:
        try:
            db.session.rollback()
        except Exception:
            pass
        flash(f"❌ Title repair failed: {e}")
    return redirect(url_for('title_repair_panel', q=q, limit=limit))

@app.route('/movie_aliases')
@login_required
@require_admin
def movie_aliases_panel():
    q = (request.args.get('q') or '').strip()
    try:
        limit = max(1, min(int(request.args.get('limit') or 50), 200))
    except Exception:
        limit = 50
    rows_q = Movie.query
    if q:
        like = f"%{q}%"
        rows_q = rows_q.filter(or_(Movie.title.ilike(like), Movie.code.ilike(like)))
    movies = rows_q.order_by(Movie.updated_at.desc().nullslast(), Movie.id.desc()).limit(limit).all()
    rows = []
    for movie in movies:
        aliases = find_movie_aliases(movie, limit=20)
        if q and not aliases and q.lower() not in (movie.title or '').lower() and q.lower() not in (movie.code or '').lower():
            continue
        rows.append({'movie': movie, 'aliases': aliases})
    return render_template('movie_aliases.html', rows=rows, q=q, limit=limit)


@app.post('/movie_aliases/add')
@login_required
@require_admin
def movie_aliases_add():
    code = (request.form.get('movie_code') or '').strip().upper()
    alias = (request.form.get('alias') or '').strip()
    q = (request.form.get('q') or '').strip()
    try:
        limit = max(1, min(int(request.form.get('limit') or 50), 200))
    except Exception:
        limit = 50
    movie = Movie.query.filter_by(code=code).first() if code else None
    if not movie:
        flash('❌ Movie not found for alias add')
        return redirect(url_for('movie_aliases_panel', q=q, limit=limit))
    try:
        result = add_movie_alias_db(movie, alias, source='web_manual')
        if result.get('changed'):
            flash(f"✅ Alias added for {movie.code}: {result.get('alias').alias}")
        elif result.get('reason') == 'conflict':
            other = result.get('movie')
            flash(f"⚠️ Alias already belongs to {(other.code if other else 'another movie')}")
        elif result.get('reason') == 'same_title':
            flash('ℹ️ Alias is the same as the movie title')
        else:
            flash('ℹ️ No alias added')
    except Exception as e:
        try:
            db.session.rollback()
        except Exception:
            pass
        flash(f"❌ Alias add failed: {e}")
    return redirect(url_for('movie_aliases_panel', q=q, limit=limit))


@app.post('/movie_aliases/delete')
@login_required
@require_admin
def movie_aliases_delete():
    raw = (request.form.get('alias_id') or '').strip()
    q = (request.form.get('q') or '').strip()
    try:
        limit = max(1, min(int(request.form.get('limit') or 50), 200))
    except Exception:
        limit = 50
    try:
        aid = int(raw)
    except Exception:
        flash('❌ Alias ID invalid')
        return redirect(url_for('movie_aliases_panel', q=q, limit=limit))
    try:
        result = delete_movie_alias_db(aid)
        if result.get('changed'):
            flash(f"✅ Alias deleted: {result.get('alias')}")
        else:
            flash('❌ Alias not found')
    except Exception as e:
        try:
            db.session.rollback()
        except Exception:
            pass
        flash(f"❌ Alias delete failed: {e}")
    return redirect(url_for('movie_aliases_panel', q=q, limit=limit))


@app.route('/resolve_tools')
@login_required
@require_admin
def resolve_tools_panel():
    q = (request.args.get('q') or '').strip()
    raw_chat_id = (request.args.get('chat_id') or '').strip()
    try:
        limit = max(1, min(int(request.args.get('limit') or 8), 50))
    except Exception:
        limit = 8
    parsed_general = parse_movie_from_filename(q) if q else None
    parsed_helper = _parse_movie_from_role_helper_filename(q) if q else None
    clean_title = _clean_movie_title_candidate(q) if q else ''
    movie = None
    matches = []
    if q:
        try:
            movie, matches = _resolve_movie_query(q)
        except Exception:
            movie, matches = None, []
    bound_movie = None
    ctx_row = None
    recent_imports = []
    chat_id = None
    if raw_chat_id:
        try:
            chat_id = int(raw_chat_id)
            bound_movie = Movie.query.filter_by(vo_group_chat_id=chat_id).first()
            ctx_row = GroupMovieContext.query.filter_by(tg_chat_id=chat_id).first()
            recent_imports = (
                GroupRoleImportRequest.query
                .filter_by(tg_chat_id=chat_id)
                .order_by(GroupRoleImportRequest.created_at.desc())
                .limit(6)
                .all()
            )
        except Exception:
            chat_id = None
    return render_template(
        'resolve_tools.html',
        q=q,
        chat_id=raw_chat_id,
        limit=limit,
        parsed_general=parsed_general,
        parsed_helper=parsed_helper,
        clean_title=clean_title,
        resolved=movie,
        matches=matches[:limit],
        bound_movie=bound_movie,
        ctx_row=ctx_row,
        recent_imports=recent_imports,
    )


def _role_import_back_args() -> dict:
    return {
        'status': (request.values.get('status') or 'PENDING').strip().upper(),
        'q': (request.values.get('q') or '').strip(),
        'limit': request.values.get('limit') or 50,
    }


def _role_import_redirect(req_id: int | None = None):
    back = _role_import_back_args()
    if (request.values.get('redirect_to') or '').strip().lower() == 'detail' and req_id:
        return redirect(url_for('role_import_detail', req_id=req_id, **back))
    return redirect(url_for('role_imports_panel', **back))


def _role_import_row(req: GroupRoleImportRequest, *, preview_limit: int = 12, raw_limit: int = 8):
    roles = _load_import_req_roles(req)
    suggestions = _load_import_req_suggestions(req)
    total_lines = sum(int(lines or 0) for _, lines in roles)
    assigned = sum(1 for row in suggestions if str(row.get('vo') or '').strip() and str(row.get('vo') or '').strip() != '(unassigned)')
    return {
        'req': req,
        'roles_count': len(roles),
        'total_lines': total_lines,
        'suggestions_count': len(suggestions),
        'assigned_count': assigned,
        'unassigned_count': max(0, len(suggestions) - assigned),
        'movie_label': f"{_movie_display_title(req.title, req.year)} — {(req.lang or '').upper() or 'BN'}",
        'existing_movie': _find_movie_by_title(req.title or '', req.year, req.lang),
        'raw_preview': [ln.strip() for ln in (req.roles_text or '').splitlines() if ln.strip()][:raw_limit],
        'raw_lines': [ln.strip() for ln in (req.roles_text or '').splitlines() if ln.strip()],
        'suggestions': suggestions[:preview_limit],
        'all_suggestions': suggestions,
        'roles': roles,
    }


def _role_import_rows(status: str = "PENDING", q: str = "", limit: int = 50):
    if not _role_import_feature_ready():
        return []
    try:
        q_obj = GroupRoleImportRequest.query
        if status and status.lower() != 'all':
            q_obj = q_obj.filter(GroupRoleImportRequest.status == status.upper())
        raw = (q or '').strip()
        if raw:
            like = f"%{raw}%"
            q_obj = q_obj.filter(
                GroupRoleImportRequest.title.ilike(like)
                | GroupRoleImportRequest.requested_by_name.ilike(like)
                | GroupRoleImportRequest.lang.ilike(like)
                | func.cast(GroupRoleImportRequest.id, db.String).ilike(like)
            )
        rows = q_obj.order_by(GroupRoleImportRequest.created_at.desc()).limit(int(limit or 50)).all()
        out=[]
        for req in rows:
            try:
                out.append(_role_import_row(req, preview_limit=12, raw_limit=8))
            except Exception:
                log.exception("Role import row render failed for req_id=%s", getattr(req, 'id', None))
        return out
    except Exception:
        try:
            db.session.rollback()
        except Exception:
            pass
        log.exception("Role imports panel query failed")
        return []


@app.route('/role_imports')
@login_required
@require_admin
def role_imports_panel():
    status = (request.args.get('status') or 'PENDING').strip().upper()
    q = (request.args.get('q') or '').strip()
    try:
        limit = max(1, min(int(request.args.get('limit') or 50), 200))
    except Exception:
        limit = 50
    if not _role_import_feature_ready():
        flash('⚠️ Role import queue is not ready yet. Redeploy once to let schema migration finish.')
        return render_template('role_imports.html', rows=[], status=status, q=q, limit=limit)
    rows = _role_import_rows(status=status, q=q, limit=limit)
    if not rows and status == 'PENDING':
        # If the route was previously crashing on a bad row or stale DB, keep the page usable.
        pass
    return render_template('role_imports.html', rows=rows, status=status, q=q, limit=limit)


@app.route('/role_imports/<int:req_id>')
@login_required
@require_admin
def role_import_detail(req_id: int):
    if not _role_import_feature_ready():
        flash('⚠️ Role import queue is not ready yet.')
        return redirect(url_for('role_imports_panel'))
    try:
        req = GroupRoleImportRequest.query.filter_by(id=req_id).first_or_404()
        row = _role_import_row(req, preview_limit=9999, raw_limit=9999)
    except Exception:
        try:
            db.session.rollback()
        except Exception:
            pass
        log.exception('Role import detail failed for req_id=%s', req_id)
        flash(f'❌ Failed to open role import #{req_id}')
        return redirect(url_for('role_imports_panel'))
    back = _role_import_back_args()
    return render_template('role_import_detail.html', row=row, **back)


@app.post('/role_imports/run')
@login_required
@require_admin
def role_imports_run():
    rid = int(request.form.get('req_id') or 0)
    action = (request.form.get('action') or '').strip().lower()
    mode = normalize_priority_mode(request.form.get('mode') or 'urgent')
    if not _role_import_feature_ready():
        flash('⚠️ Role import queue is not ready yet.')
        return _role_import_redirect(rid or None)
    req = GroupRoleImportRequest.query.filter_by(id=rid).first()
    if not req:
        flash('❌ Pending role request not found')
        return _role_import_redirect(rid or None)

    if action == 'refresh':
        if req.status != 'PENDING':
            flash(f'⚠️ Request already {req.status}; refresh only works for pending requests')
            return _role_import_redirect(req.id)
        if req.expires_at and req.expires_at < datetime.utcnow():
            req.status = 'EXPIRED'
            db.session.commit()
            flash('⏳ Request expired')
            return _role_import_redirect(req.id)
        try:
            suggestions = _refresh_role_import_request(req, commit=True)
            flash(f'🔄 Refreshed role import #{req.id} • suggestions={len(suggestions)}')
        except Exception as e:
            try:
                db.session.rollback()
            except Exception:
                pass
            flash(f'❌ Role import refresh failed: {e}')
        return _role_import_redirect(req.id)

    if req.status != 'PENDING':
        flash(f'⚠️ Request already {req.status}')
        return _role_import_redirect(req.id)
    if req.expires_at and req.expires_at < datetime.utcnow():
        req.status = 'EXPIRED'
        db.session.commit()
        flash('⏳ Request expired')
        return _role_import_redirect(req.id)
    if action == 'reject':
        req.status = 'REJECTED'
        req.reviewed_by_name = 'web'
        req.reviewed_at = datetime.utcnow()
        db.session.commit()
        flash(f'❌ Rejected role import #{req.id}')
        return _role_import_redirect(req.id)
    if action != 'approve':
        flash('❌ Unknown action')
        return _role_import_redirect(req.id)
    try:
        movie = upsert_movie(req.title, int(req.year), req.lang or 'bn')
        movie.vo_group_chat_id = int(req.tg_chat_id) if req.tg_chat_id else movie.vo_group_chat_id
        movie.updated_at = datetime.utcnow()
        db.session.commit()
        roles = _load_import_req_roles(req)
        suggestions = _auto_assign_movie_roles(
            movie,
            roles,
            urgent=priority_mode_urgent_only(mode),
            replace_existing=True,
            priority_mode=mode,
        )
        req.status = 'APPROVED'
        req.reviewed_by_name = 'web'
        req.reviewed_at = datetime.utcnow()
        db.session.commit()
        record_movie_event(movie, 'IMPORT_APPROVE', f'Web approved role import #{req.id}', detail=f'mode={mode} • assignments={len(suggestions)}', actor_source='web', actor_name='role_imports_run')
        flash(f"✅ Approved role import #{req.id} for {_movie_display_title(movie.title, movie.year)} [{movie.code}] • {priority_mode_label(mode)} ({priority_mode_hours(mode)}h)")
    except Exception as e:
        try:
            db.session.rollback()
        except Exception:
            pass
        flash(f'❌ Role import apply failed: {e}')
    return _role_import_redirect(req.id)


@app.get("/assignments")
@login_required
@require_admin
def assignments_panel():
    raw_rows = (
        db.session.query(
            Assignment.project.label("project"),
            func.count(Assignment.id).label("roles"),
            func.sum(Assignment.lines).label("total_lines"),
            func.max(Assignment.created_at).label("last_created"),
        )
        .group_by(Assignment.project)
        .order_by(func.max(Assignment.created_at).desc())
        .all()
    )

    codes = [str(r.project).strip().upper() for r in raw_rows if (r.project or "").strip()]
    movies_by_code = {}
    if codes:
        try:
            movies_by_code = {m.code: m for m in Movie.query.filter(_active_movie_expr()).filter(Movie.code.in_(codes)).all() if m.code}
        except Exception:
            try:
                db.session.rollback()
            except Exception:
                pass
            movies_by_code = {}

    rows = []
    for r in raw_rows:
        project = (r.project or "").strip()
        movie = movies_by_code.get(project.upper())
        display_name = _movie_display_title(movie.title, movie.year) if movie else project
        subtitle_bits = []
        if movie and movie.code:
            subtitle_bits.append(movie.code)
        if movie and movie.lang:
            subtitle_bits.append((movie.lang or "").upper())
        rows.append(
            {
                "project": project,
                "project_code": (movie.code if movie and movie.code else project),
                "display_name": display_name,
                "subtitle": " • ".join([x for x in subtitle_bits if x]),
                "roles": int(r.roles or 0),
                "total_lines": int(r.total_lines or 0),
                "last_created": r.last_created,
            }
        )

    return render_template("assignments.html", rows=rows)


@app.post("/assignments/new")
@login_required
@require_admin
def assignments_new():
    """Quick-start create: accept movie title OR code, auto-create movie/code, then auto-assign VOs."""
    from ops_log import log_event

    project_input = (request.form.get("project") or "").strip()
    title_input = (request.form.get("title") or "").strip()
    year = (request.form.get("year") or "").strip() or None
    lang = (request.form.get("lang") or "").strip() or None
    mode = normalize_priority_mode(request.form.get("mode") or "urgent")
    urgent = priority_mode_urgent_only(mode)
    default_deadline = priority_mode_deadline(mode)
    text = request.form.get("bulk_text") or ""

    parsed = parse_lines(text)
    if not parsed:
        flash("❌ Nothing parsed — paste role list like: man-1 120")
        return redirect(url_for("assignments_panel"))

    try:
        movie, created_movie, title_first_mode = _resolve_or_create_assignment_movie(
            project_raw=project_input,
            title_raw=title_input,
            year_raw=year,
            lang_raw=lang,
        )
    except ValueError as e:
        flash(f"❌ {e}")
        return redirect(url_for("assignments_panel"))
    except Exception as e:
        try:
            db.session.rollback()
        except Exception:
            pass
        flash(f"❌ Movie create/resolve failed: {e}")
        return redirect(url_for("assignments_panel"))

    project = (movie.code or "").strip()
    title = (movie.title or title_input or project_input or project).strip()
    lang = (movie.lang or lang or "").strip() or None

    # Aggregate duplicates (same man1 repeated) -> one bucket
    agg: dict[str, int] = {}
    for role, lines in parsed:
        agg[role] = agg.get(role, 0) + int(lines or 0)
    parsed2 = sorted(agg.items(), key=lambda x: x[0])

    # Replace assignments for this project
    try:
        Assignment.query.filter_by(project=project).delete()
        db.session.commit()
    except Exception:
        try:
            db.session.rollback()
        except Exception:
            pass

    # Workload across projects (used by pick_vo preference rules).
    rows = (
        db.session.query(Assignment.vo, func.count(func.distinct(Assignment.project)))
        .group_by(Assignment.vo)
        .all()
    )
    project_counts = {vo: int(cnt or 0) for vo, cnt in rows}

    load = movie_load(project)
    used = set()

    created = 0
    for role, lines in parsed2:
        gender = role_gender(role)
        q = VOTeam.query.filter_by(active=True, gender=gender)
        if urgent:
            q = q.filter_by(urgent_ok=True)

        picked = pick_vo(q.all(), used, load, project_counts)
        if not picked:
            continue

        used.add(picked.name)
        db.session.add(
            Assignment(
                project=project,
                movie_id=(movie.id if movie else None),
                vo=picked.name,
                role=role,
                lines=int(lines or 0),
                urgent=bool(urgent),
                priority_mode=mode,
                deadline_at=default_deadline,
            )
        )
        created += 1

    try:
        db.session.commit()
    except Exception as e:
        try:
            db.session.rollback()
        except Exception:
            pass
        flash(f"❌ Create assignments failed: {e}")
        return redirect(url_for("assignments_panel"))

    record_movie_event(movie, "CREATE_PROJECT", f"Web created {created} role(s)", detail=f"mode={mode} • urgent={urgent} • created_movie={created_movie} • lang={movie.lang or '-'}", actor_source="web", actor_name="assignments_new")
    log_event(
        "INFO",
        "web.assignments_new",
        f"Created project={project} title={title} roles={created} urgent={urgent} created_movie={created_movie}",
        traceback=json.dumps(
            {
                "project": project,
                "title": title,
                "year": movie.year,
                "lang": movie.lang,
                "roles": created,
                "urgent": urgent,
                "priority_mode": mode,
                "default_deadline_hours": priority_mode_hours(mode),
                "created_movie": created_movie,
                "title_first_mode": title_first_mode,
            },
            indent=2,
        ),
    )
    flash(f"✅ Created {created} roles for {_movie_display_title(movie.title, movie.year)} [{movie.code}]")
    return redirect(url_for("assignments_view", project=project))


@app.route("/assignments/view")
@login_required
@require_admin
def assignments_view():
    project = request.args.get("project", "").strip()
    if not project:
        return redirect(url_for("assignments_panel"))

    movie = Movie.query.filter_by(code=project).first()
    if not movie and not _looks_like_movie_code(project):
        guessed = _find_movie_by_title(project)
        if guessed:
            return redirect(url_for("assignments_view", project=guessed.code))

    items = Assignment.query.filter_by(project=project).order_by(Assignment.role.asc()).all()

    # Submission lookup for DONE/PENDING
    try:
        subs = VORoleSubmission.query.filter_by(movie=project).with_entities(
            VORoleSubmission.vo, VORoleSubmission.role
        ).all()
    except Exception as e:
        log.warning("assignments_view: VORoleSubmission query failed (empty done_set): %s", e)
        try:
            db.session.rollback()
        except Exception:
            pass
        subs = []
    done_set = {(v, r) for (v, r) in subs}

    # Enrich items with done flag (safe attribute)
    for a in items:
        a.done = (a.vo, a.role) in done_set

    # VO summary: total roles + pending roles
    vo_totals = {}
    for a in items:
        s = vo_totals.setdefault(a.vo, {"total": 0, "pending": 0, "lines": 0})
        s["total"] += 1
        s["lines"] += int(a.lines or 0)
        if not a.done:
            s["pending"] += 1

    vo_summary = sorted(
        [(vo, d["total"], d["pending"], d["lines"]) for vo, d in vo_totals.items()],
        key=lambda x: (-x[2], -x[1], x[0].lower()),
    )

    project_label = _movie_display_title(movie.title, movie.year) if movie else project
    return render_template(
        "assignment_view.html",
        project=project,
        project_label=project_label,
        movie=movie,
        items=items,
        vo_summary=vo_summary,
    )


@app.route("/movie_history")
@login_required
@require_admin
def movie_history_panel():
    project = (request.args.get("project") or request.args.get("code") or "").strip()
    if not project:
        flash("❌ Missing movie code/title for history")
        return redirect(url_for("assignments_panel"))

    movie = Movie.query.filter_by(code=project).first()
    if not movie and not _looks_like_movie_code(project):
        guessed = _find_movie_by_title(project)
        if not guessed:
            guessed = Movie.query.filter(func.lower(Movie.title) == project.lower()).order_by(Movie.id.desc()).first()
        if guessed:
            movie = guessed
    if not movie:
        movie = Movie.query.filter_by(code=project).first()
    if not movie:
        flash(f"❌ Movie not found for history: {project}")
        return redirect(url_for("assignments_panel"))

    rows = []
    for ev in fetch_movie_history(movie, limit=80):
        rows.append({
            "ts": ev.created_at.strftime("%Y-%m-%d %H:%M") if ev.created_at else "-",
            "event_type": ev.event_type or "INFO",
            "summary": ev.summary or "",
            "detail": ev.detail or "",
            "actor": " / ".join([x for x in [ev.actor_source, ev.actor_name] if x]) or "-",
        })
    return render_template("movie_history.html", movie=movie, rows=rows)


def _activity_filter_state():
    q = (request.args.get("q") or "").strip()
    event_type = (request.args.get("event_type") or "all").strip().upper() or "all"
    source = (request.args.get("source") or "all").strip().lower() or "all"
    archived = (request.args.get("archived") or "all").strip().lower() or "all"
    try:
        limit = int(request.args.get("limit") or 120)
    except Exception:
        limit = 120
    limit = max(10, min(limit, 500))
    return q, event_type, source, archived, limit


def _activity_rows(q: str = "", event_type: str = "all", source: str = "all", archived: str = "all", limit: int = 120):
    rows = []
    events = fetch_recent_movie_events(limit=limit, include_archived=True)
    codes = {str((ev.movie_code or "")).strip() for ev in events if (ev.movie_code or "").strip()}
    archived_map = {}
    if codes:
        for m in Movie.query.filter(Movie.code.in_(list(codes))).all():
            archived_map[str((m.code or "")).strip()] = bool(getattr(m, "is_archived", False))
    qq = (q or "").lower().strip()
    event_type = (event_type or "all").upper()
    source = (source or "all").lower()
    archived = (archived or "all").lower()
    for ev in events:
        code = (ev.movie_code or "").strip()
        is_archived = bool(archived_map.get(code, False)) if code else False
        hay = " | ".join([
            (ev.movie_code or ""),
            (ev.movie_title or ""),
            (ev.summary or ""),
            (ev.detail or ""),
            (ev.event_type or ""),
            (ev.actor_source or ""),
            (ev.actor_name or ""),
        ]).lower()
        if qq and qq not in hay:
            continue
        if event_type != "ALL" and (ev.event_type or "").upper() != event_type:
            continue
        if source != "all" and (ev.actor_source or "").lower() != source:
            continue
        if archived == "archived" and not is_archived:
            continue
        if archived == "active" and is_archived:
            continue
        rows.append({
            "ts": ev.created_at.strftime("%Y-%m-%d %H:%M") if ev.created_at else "-",
            "movie_code": code or "-",
            "movie_title": (ev.movie_title or "").strip() or (code or "-"),
            "event_type": ev.event_type or "INFO",
            "summary": ev.summary or "",
            "detail": ev.detail or "",
            "actor": " / ".join([x for x in [ev.actor_source, ev.actor_name] if x]) or "-",
            "actor_source": (ev.actor_source or "").strip() or "-",
            "actor_name": (ev.actor_name or "").strip() or "-",
            "is_archived": is_archived,
        })
    return rows


def _bulk_movies_filter_state():
    q = (request.args.get("q") or "").strip()
    scope = (request.args.get("scope") or "active").strip().lower() or "active"
    preset = (request.args.get("preset") or "").strip().lower()
    if scope not in {"active", "archived"}:
        scope = "active"
    try:
        limit = int(request.args.get("limit") or 50)
    except Exception:
        limit = 50
    limit = max(5, min(limit, 200))
    return q, scope, limit, preset


def _bulk_saved_filter_slots():
    slots = []
    for idx in range(1, 4):
        raw = (kv_get(f"bulk_filter_slot_{idx}") or "").strip()
        label = (kv_get(f"bulk_filter_slot_{idx}_label") or f"Slot {idx}").strip() or f"Slot {idx}"
        state = {}
        if raw:
            try:
                state = json.loads(raw)
            except Exception:
                state = {}
        q = str(state.get("q") or "").strip()
        scope = str(state.get("scope") or "active").strip().lower() or "active"
        if scope not in {"active", "archived"}:
            scope = "active"
        preset = str(state.get("preset") or "").strip().lower()
        try:
            limit = int(state.get("limit") or 50)
        except Exception:
            limit = 50
        limit = max(5, min(limit, 200))
        slots.append({
            "slot": idx,
            "label": label,
            "has_value": bool(raw),
            "q": q,
            "scope": scope,
            "limit": limit,
            "preset": preset,
        })
    return slots


def _bulk_save_filter_slot(slot: int, q: str, scope: str, limit: int, preset: str = "", label: str = "") -> bool:
    slot = int(slot)
    payload = {
        "q": (q or "").strip(),
        "scope": (scope or "active").strip().lower() or "active",
        "limit": int(limit or 50),
        "preset": (preset or "").strip().lower(),
    }
    ok1 = kv_set(f"bulk_filter_slot_{slot}", json.dumps(payload))
    clean_label = (label or f"Slot {slot}").strip() or f"Slot {slot}"
    ok2 = kv_set(f"bulk_filter_slot_{slot}_label", clean_label)
    return bool(ok1 and ok2)


def _cleanup_preset_specs():
    return [
        {
            "key": "inactive14",
            "scope": "active",
            "title": "Inactive 14 days",
            "desc": "Active movies with no translator and no active roles, untouched for 14+ days.",
        },
        {
            "key": "inactive30",
            "scope": "active",
            "title": "Inactive 30 days",
            "desc": "Stricter version for older inactive movies.",
        },
        {
            "key": "no_tr_no_roles",
            "scope": "active",
            "title": "No translator + no roles",
            "desc": "Active movies that still have no translator and no VO role rows.",
        },
        {
            "key": "translator_only14",
            "scope": "active",
            "title": "Translator only, no VO roles (14d)",
            "desc": "Movies with translator/task presence but no active VO roles for 14+ days.",
        },
        {
            "key": "archived14",
            "scope": "archived",
            "title": "Archived 14 days",
            "desc": "Archived movies older than 14 days, ready for review or cleanup.",
        },
        {
            "key": "archived30",
            "scope": "archived",
            "title": "Archived 30 days",
            "desc": "Archived movies older than 30 days, best candidates for permanent cleanup.",
        },
    ]


def _row_matches_bulk_preset(row: dict, preset: str) -> bool:
    preset = (preset or "").strip().lower()
    if not preset:
        return True
    age_days = int(row.get("age_days") or 0)
    roles = int(row.get("roles") or 0)
    has_tr = bool((row.get("translator_name") or "").strip()) or bool(row.get("has_translation_task"))
    is_archived = bool(row.get("is_archived"))
    if preset == "inactive14":
        return (not is_archived) and roles == 0 and (not has_tr) and age_days >= 14
    if preset == "inactive30":
        return (not is_archived) and roles == 0 and (not has_tr) and age_days >= 30
    if preset == "no_tr_no_roles":
        return (not is_archived) and roles == 0 and (not has_tr)
    if preset == "translator_only14":
        return (not is_archived) and roles == 0 and has_tr and age_days >= 14
    if preset == "archived14":
        return is_archived and age_days >= 14
    if preset == "archived30":
        return is_archived and age_days >= 30
    return True


def _bulk_movie_rows(q: str = "", scope: str = "active", limit: int = 50, preset: str = ""):
    scope = (scope or "active").lower()
    preset = (preset or "").strip().lower()
    rows_q = Movie.query.filter(_archived_movie_expr() if scope == "archived" else _active_movie_expr())
    if q:
        code = q.strip().upper()
        rows_q = rows_q.filter(or_(Movie.code == code, Movie.title.ilike(f"%{q}%")))
    fetch_limit = 500 if preset else limit
    movies = rows_q.order_by(Movie.updated_at.desc().nullslast(), Movie.id.desc()).limit(fetch_limit).all()
    codes = [str((m.code or "")).strip() for m in movies if (m.code or "").strip()]
    mids = [m.id for m in movies if getattr(m, "id", None)]

    role_counts = {}
    line_counts = {}
    if codes or mids:
        try:
            role_rows = (
                db.session.query(Assignment.project, func.count(Assignment.id), func.coalesce(func.sum(Assignment.lines), 0))
                .filter(or_(Assignment.project.in_(codes or ["__none__"]), Assignment.movie_id.in_(mids or [-1])))
                .group_by(Assignment.project)
                .all()
            )
            for project, cnt, total_lines in role_rows:
                key = str(project or "").strip()
                role_counts[key] = int(cnt or 0)
                line_counts[key] = int(total_lines or 0)
        except Exception:
            try:
                db.session.rollback()
            except Exception:
                pass

    task_map = {}
    if codes or mids:
        try:
            tasks = (
                TranslationTask.query
                .filter(or_(TranslationTask.movie_code.in_(codes or ["__none__"]), TranslationTask.movie_id.in_(mids or [-1])))
                .order_by(TranslationTask.updated_at.desc().nullslast(), TranslationTask.id.desc())
                .all()
            )
            for t in tasks:
                key = str((t.movie_code or "")).strip()
                if not key and getattr(t, "movie_id", None):
                    for m in movies:
                        if m.id == t.movie_id:
                            key = str((m.code or "")).strip()
                            break
                if key and key not in task_map:
                    task_map[key] = t
        except Exception:
            try:
                db.session.rollback()
            except Exception:
                pass

    rows = []
    for m in movies:
        code = str((m.code or "")).strip()
        task = task_map.get(code)
        subtitle_bits = []
        if m.year:
            subtitle_bits.append(str(m.year))
        if m.lang:
            subtitle_bits.append((m.lang or "").upper())
        if m.translator_assigned:
            subtitle_bits.append(f"TR {m.translator_assigned}")
        elif task and getattr(task, "translator_name", None):
            subtitle_bits.append(f"TR {task.translator_name}")
        updated_dt = (m.updated_at or m.created_at)
        translator_name = (m.translator_assigned or "").strip() or (getattr(task, "translator_name", None) or "")
        task_status = (getattr(task, "status", None) or "").strip().upper()
        age_days = 0
        if updated_dt:
            try:
                age_days = max(0, (datetime.utcnow() - updated_dt).days)
            except Exception:
                age_days = 0
        rows.append({
            "code": code or "-",
            "display_name": _movie_display_title(m.title, m.year),
            "subtitle": " • ".join([x for x in subtitle_bits if x]),
            "status": m.status or ("ARCHIVED" if getattr(m, "is_archived", False) else "RECEIVED"),
            "roles": int(role_counts.get(code, 0)),
            "lines": int(line_counts.get(code, 0)),
            "updated_at": updated_dt,
            "is_archived": bool(getattr(m, "is_archived", False)),
            "translator_name": translator_name,
            "task_status": task_status,
            "has_translation_task": bool(task),
            "age_days": age_days,
        })
    if preset:
        rows = [r for r in rows if _row_matches_bulk_preset(r, preset)]
    return rows[:limit]


@app.route("/bulk_movies")
@login_required
@require_admin
def bulk_movies_panel():
    q, scope, limit, preset = _bulk_movies_filter_state()
    rows = _bulk_movie_rows(q=q, scope=scope, limit=limit, preset=preset)
    preset_specs = []
    for spec in _cleanup_preset_specs():
        try:
            count = len(_bulk_movie_rows(q="", scope=spec["scope"], limit=120, preset=spec["key"]))
        except Exception:
            count = 0
        row = dict(spec)
        row["count"] = count
        preset_specs.append(row)
    return render_template(
        "bulk_movies.html",
        rows=rows,
        q=q,
        scope=scope,
        limit=limit,
        preset=preset,
        saved_slots=_bulk_saved_filter_slots(),
        preset_specs=preset_specs,
    )


@app.post("/bulk_movies/save_filter")
@login_required
@require_admin
def bulk_movies_save_filter():
    try:
        slot = int(request.form.get("slot") or 0)
    except Exception:
        slot = 0
    if slot not in {1, 2, 3}:
        flash("❌ Invalid filter slot.")
        return redirect(url_for("bulk_movies_panel"))
    q = (request.form.get("q") or "").strip()
    scope = (request.form.get("scope") or "active").strip().lower() or "active"
    preset = (request.form.get("preset") or "").strip().lower()
    label = (request.form.get("label") or "").strip()
    try:
        limit = int(request.form.get("limit") or 50)
    except Exception:
        limit = 50
    limit = max(5, min(limit, 200))
    if _bulk_save_filter_slot(slot, q=q, scope=scope, limit=limit, preset=preset, label=label):
        flash(f"✅ Saved current bulk filter to slot {slot}")
    else:
        flash("❌ Failed to save bulk filter slot")
    return redirect(url_for("bulk_movies_panel", q=q, scope=scope, limit=limit, preset=preset))


@app.route("/cleanup_presets")
@login_required
@require_admin
def cleanup_presets_panel():
    specs = []
    for spec in _cleanup_preset_specs():
        rows = _bulk_movie_rows(q="", scope=spec["scope"], limit=120, preset=spec["key"])
        entry = dict(spec)
        entry["count"] = len(rows)
        entry["sample"] = rows[:5]
        specs.append(entry)
    return render_template("cleanup_presets.html", specs=specs)


@app.post("/bulk_movies/run")
@login_required
@require_admin
def bulk_movies_run():
    action = (request.form.get("action") or "").strip().lower()
    codes = [str(x).strip().upper() for x in request.form.getlist("codes") if str(x).strip()]
    q = (request.form.get("q") or "").strip()
    scope = (request.form.get("scope") or "active").strip().lower() or "active"
    preset = (request.form.get("preset") or "").strip().lower()
    limit = request.form.get("limit") or "50"
    confirm = (request.form.get("confirm") or "").strip().upper()

    if not codes:
        flash("❌ Select at least one movie first.")
        return redirect(url_for("bulk_movies_panel", q=q, scope=scope, limit=limit, preset=preset))

    count = 0
    skipped = []
    for code in codes:
        movie = Movie.query.filter_by(code=code).first()
        if not movie:
            skipped.append(f"{code}: missing")
            continue
        if action == "archive":
            if getattr(movie, "is_archived", False):
                skipped.append(f"{code}: already archived")
                continue
            _archive_movie_record(movie, clear_active_rows=True)
            record_movie_event(movie, "ARCHIVE", "Web bulk archived movie", detail="Bulk Ops page", actor_source="web", actor_name="bulk_movies_archive")
            count += 1
        elif action == "clear":
            affected = Assignment.query.filter((Assignment.movie_id == movie.id) | (Assignment.project == code)).delete(synchronize_session=False)
            record_movie_event(movie, "CLEAR_ACTIVE", "Web bulk cleared active roles", detail=f"rows={affected}", actor_source="web", actor_name="bulk_movies_clear")
            count += 1
        elif action == "unarchive":
            if not getattr(movie, "is_archived", False):
                skipped.append(f"{code}: already active")
                continue
            movie.is_archived = False
            movie.archived_at = None
            if (movie.status or "").upper() == "ARCHIVED":
                movie.status = "RECEIVED"
            movie.updated_at = datetime.utcnow()
            record_movie_event(movie, "UNARCHIVE", "Web bulk unarchived movie", detail="Bulk Ops page", actor_source="web", actor_name="bulk_movies_unarchive")
            count += 1
        elif action == "hard_delete":
            if confirm != "DELETE":
                flash("❌ Hard delete blocked. Type DELETE in the confirmation box first.")
                return redirect(url_for("bulk_movies_panel", q=q, scope=scope, limit=limit, preset=preset))
            if not getattr(movie, "is_archived", False):
                skipped.append(f"{code}: archive first")
                continue
            record_movie_event(movie, "HARD_DELETE", "Web bulk hard deleted movie", detail="Bulk Ops page permanent delete", actor_source="web", actor_name="bulk_movies_hard_delete")
            _hard_delete_movie_record(movie)
            count += 1
        else:
            flash(f"❌ Unsupported bulk action: {action}")
            return redirect(url_for("bulk_movies_panel", q=q, scope=scope, limit=limit, preset=preset))

    db.session.commit()
    msg_map = {
        "archive": "archived",
        "clear": "cleared active roles for",
        "unarchive": "unarchived",
        "hard_delete": "hard deleted",
    }
    flash(f"✅ Bulk {msg_map.get(action, action)} {count} movie(s)")
    if skipped:
        flash("ℹ️ Skipped: " + "; ".join(skipped[:8]) + (" ..." if len(skipped) > 8 else ""))
    next_scope = "archived" if action == "archive" else ("active" if action == "unarchive" else scope)
    return redirect(url_for("bulk_movies_panel", q=q, scope=next_scope, limit=limit, preset=preset))


@app.route("/activity")
@login_required
@require_admin
def activity_panel():
    q, event_type, source, archived, limit = _activity_filter_state()
    rows = []
    try:
        rows = _activity_rows(q=q, event_type=event_type, source=source, archived=archived, limit=limit)
    except Exception as e:
        log.warning("activity_panel failed: %s", e)
        try:
            db.session.rollback()
        except Exception:
            pass
    event_types = ["ALL"] + sorted({str((r.get("event_type") or "INFO")).upper() for r in rows})
    return render_template(
        "activity.html",
        rows=rows,
        q=q,
        event_type=event_type,
        source=source,
        archived=archived,
        limit=limit,
        event_types=event_types,
    )


@app.route("/activity/export")
@login_required
@require_admin
def activity_export():
    q, event_type, source, archived, limit = _activity_filter_state()
    try:
        rows = _activity_rows(q=q, event_type=event_type, source=source, archived=archived, limit=limit)
    except Exception as e:
        log.warning("activity_export failed: %s", e)
        try:
            db.session.rollback()
        except Exception:
            pass
        rows = []
    sio = StringIO()
    writer = csv.writer(sio)
    writer.writerow(["time", "movie_code", "movie_title", "event_type", "summary", "detail", "actor_source", "actor_name", "archived"])
    for r in rows:
        actor_source, _, actor_name = str(r.get("actor") or "-").partition(" / ")
        writer.writerow([
            r.get("ts") or "",
            r.get("movie_code") or "",
            r.get("movie_title") or "",
            r.get("event_type") or "",
            r.get("summary") or "",
            r.get("detail") or "",
            r.get("actor_source") or actor_source,
            r.get("actor_name") or actor_name,
            "yes" if r.get("is_archived") else "no",
        ])
    filename = f"movie_activity_{datetime.utcnow().strftime('%Y%m%d-%H%M%S')}.csv"
    return Response(
        sio.getvalue(),
        mimetype="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.route("/assignments/clear", methods=["POST"])
@login_required
@require_admin
def assignments_clear():
    project = request.form.get("project", "").strip()
    movie = Movie.query.filter_by(code=project).first()

    if movie:
        affected = Assignment.query.filter((Assignment.movie_id == movie.id) | (Assignment.project == project)).delete(synchronize_session=False)
        record_movie_event(movie, "CLEAR_ACTIVE", f"Web cleared active roles only", detail=f"rows={affected}", actor_source="web", actor_name="assignments_clear")
    else:
        Assignment.query.filter_by(project=project).delete()
    db.session.commit()
    flash(f"✅ Cleared active assignments for {project}")
    return redirect(url_for("assignments_panel"))


@app.route("/assignments/archive", methods=["POST"])
@login_required
@require_admin
def assignments_archive():
    project = (request.form.get("project") or "").strip()
    movie = Movie.query.filter_by(code=project).first()
    if not movie:
        flash(f"❌ Movie not found for archive: {project}")
        return redirect(url_for("assignments_panel"))
    _archive_movie_record(movie, clear_active_rows=True)
    record_movie_event(movie, "ARCHIVE", "Web archived movie", detail="Hidden from Telegram search", actor_source="web", actor_name="assignments_archive")
    db.session.commit()
    flash(f"✅ Archived {_movie_display_title(movie.title, movie.year)} [{movie.code}] and hid it from Telegram search")
    return redirect(url_for("assignments_panel"))


@app.route("/assignments/unarchive", methods=["POST"])
@login_required
@require_admin
def assignments_unarchive():
    project = (request.form.get("project") or "").strip()
    movie = Movie.query.filter_by(code=project).first()
    if not movie:
        flash(f"❌ Movie not found for unarchive: {project}")
        return redirect(url_for("archived_panel"))
    if not getattr(movie, "is_archived", False):
        flash(f"ℹ️ Movie is already active: {project}")
        return redirect(url_for("assignments_view", project=project))
    movie.is_archived = False
    movie.archived_at = None
    if (movie.status or "").upper() == "ARCHIVED":
        movie.status = "RECEIVED"
    movie.updated_at = datetime.utcnow()
    record_movie_event(movie, "UNARCHIVE", "Web unarchived movie", detail="Visible in Telegram search again", actor_source="web", actor_name="assignments_unarchive")
    db.session.commit()
    flash(f"✅ Unarchived {_movie_display_title(movie.title, movie.year)} [{movie.code}]")
    return redirect(url_for("assignments_view", project=project))


@app.route("/assignments/hard_delete", methods=["POST"])
@login_required
@require_admin
def assignments_hard_delete():
    project = (request.form.get("project") or "").strip()
    confirm = (request.form.get("confirm") or "").strip().upper()
    movie = Movie.query.filter_by(code=project).first()
    if not movie:
        flash(f"❌ Movie not found for hard delete: {project}")
        return redirect(url_for("assignments_panel"))
    if confirm != "DELETE":
        flash("❌ Hard delete blocked. Type DELETE in the confirmation box first.")
        return redirect(url_for("assignments_view", project=project))
    label = _movie_display_title(movie.title, movie.year)
    code = movie.code
    record_movie_event(movie, "HARD_DELETE", "Web hard deleted movie", detail="Permanent delete requested from web", actor_source="web", actor_name="assignments_hard_delete")
    _hard_delete_movie_record(movie)
    db.session.commit()
    flash(f"✅ Hard deleted {label} [{code}] permanently")
    return redirect(url_for("assignments_panel"))


@app.route("/assignments/export")
@login_required
@require_admin
def assignments_export():
    project = request.args.get("project", "").strip()
    items = Assignment.query.filter_by(project=project).order_by(Assignment.created_at.asc()).all()

    def gen():
        yield "project,role,vo,lines,created_at\n"
        for a in items:
            yield f"{project},{a.role},{a.vo},{a.lines},{a.created_at}\n"

    return Response(
        gen(),
        mimetype="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{project}.assignments.csv"'},
    )



@app.route("/duplicates")
@login_required
@require_admin
def duplicates_panel():
    q = (request.args.get("q") or "").strip()
    try:
        limit = max(5, min(int(request.args.get("limit") or 24), 80))
    except Exception:
        limit = 24
    groups = duplicate_groups(q=q, limit=limit, include_archived=True)
    prepared = []
    for g in groups:
        items = []
        target = g.get("target")
        for m in g.get("items") or []:
            info = {
                "movie": m,
                "code": m.code,
                "display_name": _movie_display_title(m.title, m.year),
                "status": m.status or ("ARCHIVED" if getattr(m, "is_archived", False) else "RECEIVED"),
                "is_target": bool(target and m.id == target.id),
                "is_archived": bool(getattr(m, "is_archived", False)),
                "preview": merge_preview(m, target) if (target and m.id != target.id) else None,
            }
            items.append(info)
        prepared.append({
            "title": _movie_display_title(g.get("title"), g.get("year")),
            "lang": (g.get("lang") or "").upper() or "-",
            "target": target,
            "items": items,
            "count": g.get("count") or len(items),
        })
    return render_template("duplicates.html", groups=prepared, q=q, limit=limit)




@app.route("/merge_simulator")
@login_required
@require_admin
def merge_simulator_panel():
    source_code = (request.args.get("source_code") or "").strip().upper()
    target_code = (request.args.get("target_code") or "").strip().upper()
    q = (request.args.get("q") or "").strip()
    try:
        limit = max(5, min(int(request.args.get("limit") or 24), 80))
    except Exception:
        limit = 24
    if not source_code or not target_code:
        flash("❌ Source and target code are required for merge simulator.")
        return redirect(url_for("duplicates_panel", q=q, limit=limit))
    source = Movie.query.filter_by(code=source_code).first()
    target = Movie.query.filter_by(code=target_code).first()
    if not source or not target:
        flash("❌ Source or target movie not found.")
        return redirect(url_for("duplicates_panel", q=q, limit=limit))
    if source.id == target.id:
        flash("❌ Source and target cannot be the same movie.")
        return redirect(url_for("duplicates_panel", q=q, limit=limit))
    sim = merge_simulation(source, target)
    return render_template("merge_simulator.html",
        source=source,
        target=target,
        sim=sim,
        q=q,
        limit=limit,
        title="Merge Simulator",
    )

@app.post("/duplicates/merge")
@login_required
@require_admin
def duplicates_merge():
    source_code = (request.form.get("source_code") or "").strip().upper()
    target_code = (request.form.get("target_code") or "").strip().upper()
    mode = (request.form.get("mode") or "archive").strip().lower()
    confirm = (request.form.get("confirm") or "").strip().upper()
    q = (request.form.get("q") or "").strip()
    try:
        limit = max(5, min(int(request.form.get("limit") or 24), 80))
    except Exception:
        limit = 24
    if confirm != "MERGE":
        flash("❌ Merge blocked. Type MERGE in the confirmation box first.")
        return redirect(url_for("duplicates_panel", q=q, limit=limit))
    source = Movie.query.filter_by(code=source_code).first()
    target = Movie.query.filter_by(code=target_code).first()
    if not source or not target:
        flash("❌ Source or target movie not found.")
        return redirect(url_for("duplicates_panel", q=q, limit=limit))
    if source.id == target.id:
        flash("❌ Source and target cannot be the same movie.")
        return redirect(url_for("duplicates_panel", q=q, limit=limit))
    result = merge_movies(source, target, actor_source="web", actor_name="duplicates_merge", delete_source=(mode == "delete"))
    db.session.commit()
    moved = result.get("moved") or {}
    flash(f"✅ Merged {source_code} into {target_code} • moved {moved.get('total_rows', 0)} related row(s) • warnings {moved.get('warnings_count', 0)} • risk {str(moved.get('severity', 'low')).upper()} • source {result.get('source_state')}")
    return redirect(url_for("assignments_view", project=target_code))



# --------------------------------------------------
# EXPORTS / BACKUPS
# --------------------------------------------------
@app.get("/export/excel")
@login_required
@require_admin
def export_excel_full():
    """Download a full Excel export.

    Priority: NEVER crash even if schema drift exists. We export whatever tables/columns
    currently exist in the DB.
    """
    from ops_log import log_event

    try:
        xlsx_bytes, report = export_excel_dynamic(db.engine)
        log_event(
            "INFO",
            "web.export_excel",
            f"Exported Excel OK tables={len(report.get('tables', []))} errors={len(report.get('errors', []))}",
            traceback=json.dumps(report, ensure_ascii=False, indent=2),
        )
    except Exception as e:
        try:
            db.session.rollback()
        except Exception:
            pass
        # fail-soft: return an Excel with error info only
        log_event("ERROR", "web.export_excel", f"Export Excel FAILED: {e}")
        wb = Workbook()
        ws = wb.active
        ws.title = "Export_Error"
        ws["A1"] = "error"
        ws["B1"] = str(e)
        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        xlsx_bytes = bio.read()

    fname = f"vo_tracker_export_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}Z.xlsx"
    return send_file(
        BytesIO(xlsx_bytes),
        as_attachment=True,
        download_name=fname,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _model_rows(model, exclude: set[str] | None = None):
    exclude = set(exclude or set())
    out = []
    for row in model.query.all():
        d = {}
        for col in row.__table__.columns:
            k = col.name
            if k in exclude:
                continue
            v = getattr(row, k)
            if isinstance(v, datetime):
                try:
                    v = v.replace(tzinfo=timezone.utc).isoformat()
                except Exception:
                    v = v.isoformat() if hasattr(v, "isoformat") else str(v)
            d[k] = v
        out.append(d)
    return out


@app.get("/backup/json")
@login_required
@require_admin
def backup_json_zip():
    """Download a zipped JSON backup.

    Priority: NEVER crash even if schema drift exists. We export whatever tables/columns
    currently exist in the DB.
    """
    from ops_log import log_event

    stamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    try:
        zip_bytes, report = backup_json_zip_dynamic(db.engine, app_version=APP_VERSION)
        log_event(
            "INFO",
            "web.backup_json",
            f"Backup JSON OK tables={len(report.get('tables', []))} errors={len(report.get('errors', []))}",
            traceback=json.dumps(report, ensure_ascii=False, indent=2),
        )
    except Exception as e:
        try:
            db.session.rollback()
        except Exception:
            pass
        log_event("ERROR", "web.backup_json", f"Backup JSON FAILED: {e}")
        # Return a zip containing only meta.json with error.
        report = {
            "exported_at": datetime.utcnow().replace(tzinfo=timezone.utc).isoformat(),
            "app_version": APP_VERSION,
            "tables": [],
            "errors": [{"error": str(e)}],
        }
        zbio = BytesIO()
        with zipfile.ZipFile(zbio, "w", compression=zipfile.ZIP_DEFLATED) as z:
            z.writestr("meta.json", json.dumps(report, ensure_ascii=False, indent=2))
        zbio.seek(0)
        zip_bytes = zbio.read()

    return send_file(
        BytesIO(zip_bytes),
        as_attachment=True,
        download_name=f"vo_tracker_backup_{stamp}Z.zip",
        mimetype="application/zip",
    )




# --------------------------------------------------
# BACKUP TO TELEGRAM (manual + cron trigger)
# --------------------------------------------------
def _tg_token_tail(tok: str) -> str:
    if not tok:
        return ""
    return tok[-6:] if len(tok) >= 6 else tok

def _make_backup_artifacts():
    """Generate backup artifacts (Excel, JSON ZIP, logs.txt) as bytes.
    Never raises: returns (artifacts_dict, report_dict).
    """
    from ops_log import log_event

    report = {"excel": None, "json_zip": None, "logs_txt": None, "errors": []}
    artifacts = {}
    # Excel
    try:
        xlsx_bytes, rep = export_excel_dynamic(db.engine)
        artifacts["excel"] = ("vo_tracker_export_%sZ.xlsx" % datetime.utcnow().strftime("%Y%m%d_%H%M%S"), xlsx_bytes)
        report["excel"] = {"tables": rep.get("tables", []), "errors": rep.get("errors", [])}
    except Exception as e:
        try:
            db.session.rollback()
        except Exception:
            pass
        report["errors"].append({"artifact": "excel", "error": str(e)})
        log_event("ERROR", "backup.make", f"Excel generate failed: {e}")

    # JSON ZIP
    try:
        zbytes, rep = backup_json_zip_dynamic(db.engine, app_version=APP_VERSION)
        artifacts["json_zip"] = ("vo_tracker_backup_%sZ.zip" % datetime.utcnow().strftime("%Y%m%d_%H%M%S"), zbytes)
        report["json_zip"] = {"tables": rep.get("tables", []), "errors": rep.get("errors", [])}
    except Exception as e:
        try:
            db.session.rollback()
        except Exception:
            pass
        report["errors"].append({"artifact": "json_zip", "error": str(e)})

    # logs.txt
    try:
        limit = int(os.getenv("EXPORT_MAX_LOGS", "5000"))
    except Exception:
        limit = 5000
    limit = max(1, min(limit, 50000))
    try:
        items = fetch_logs(limit=limit)
        out_lines = []
        for it in items:
            out_lines.append(f"[{it.get('ts','')}] {it.get('level','INFO')} {it.get('source','')}: {it.get('message','')}")
            tb = (it.get("traceback") or "").strip()
            if tb:
                out_lines.append(tb)
                out_lines.append("")
        logs_txt = ("\n".join(out_lines) + "\n").encode("utf-8", errors="ignore")
        artifacts["logs_txt"] = ("logs_%sZ.txt" % datetime.utcnow().strftime("%Y%m%d_%H%M%S"), logs_txt)
    except Exception as e:
        try:
            db.session.rollback()
        except Exception:
            pass
        report["errors"].append({"artifact": "logs_txt", "error": str(e)})

    report["generated_at"] = datetime.utcnow().replace(tzinfo=timezone.utc).isoformat()
    report["app_version"] = APP_VERSION
    return artifacts, report


def _tg_send_document(chat_id: str, filename: str, data_bytes: bytes, caption: str | None = None) -> dict:
    """Send a document to Telegram via Bot API. Returns Telegram JSON payload (result)."""
    import httpx

    if not BOT_TOKEN:
        raise RuntimeError("BOT_TOKEN is not set")
    url = f"https://api.telegram.org/bot{BOT_TOKEN}/sendDocument"
    form = {"chat_id": chat_id}
    if caption:
        form["caption"] = caption
        form["parse_mode"] = "HTML"
    files = {"document": (filename, data_bytes)}
    r = httpx.post(url, data=form, files=files, timeout=90)
    r.raise_for_status()
    payload = r.json()
    if not payload.get("ok"):
        raise RuntimeError(str(payload))
    return payload.get("result") or payload


@app.post("/backups/send_telegram")
@login_required
@require_admin
def backups_send_telegram():
    """Generate and send backups to Telegram chat configured by BACKUP_TELEGRAM_CHAT_ID or form field."""
    from ops_log import log_event

    chat_id = get_backup_chat_id(request.form.get("chat_id"))
    if not chat_id:
        flash("BACKUP_TELEGRAM_CHAT_ID not set. Set env var or enter Chat ID here.", "error")
        return redirect(url_for("backups_panel"))

    kinds = set(request.form.getlist("kinds") or ["all"])
    # normalize
    if "all" in kinds:
        kinds = {"excel", "json_zip", "logs_txt"}

    artifacts, report = _make_backup_artifacts()
    sent = []
    errors = []

    caption_header = f"VO Tracker Backup • {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC • v{APP_VERSION}"
    for k in ["excel", "json_zip", "logs_txt"]:
        if k not in kinds:
            continue
        if k not in artifacts:
            continue
        fname, bts = artifacts[k]
        try:
            res = _tg_send_document(chat_id, fname, bts, caption=caption_header)
            sent.append({"kind": k, "file": fname, "message_id": res.get("message_id")})
        except Exception as e:
            errors.append({"kind": k, "file": fname, "error": str(e)})

    report["sent"] = sent
    report["send_errors"] = errors
    report["chat_id"] = chat_id

    level = "INFO" if not errors else "WARN"
    log_event(
        level,
        "web.backup_send_tg",
        f"Sent to TG chat={chat_id} sent={len(sent)} errors={len(errors)} token_tail={_tg_token_tail(BOT_TOKEN)}",
        traceback=json.dumps(report, ensure_ascii=False, indent=2),
    )

    if errors:
        flash(f"Backup send completed with errors: {len(errors)}. Check Logs for details.", "error")
    else:
        flash(f"Backup sent to Telegram successfully ({len(sent)} file(s)).", "success")

    return redirect(url_for("backups_panel"))

@app.post("/backups/save_destination")
@login_required
@require_admin
def backups_save_destination():
    """Save backup destination chat id into DB (app_kv)."""
    from ops_log import log_event

    chat_id = (request.form.get("chat_id") or "").strip()
    if not chat_id:
        flash("Please enter a Chat ID to save.", "error")
        return redirect(url_for("backups_panel"))

    ok = kv_set("backup_chat_id", chat_id)
    if ok:
        log_event("INFO", "web.backup_dest_set", f"Backup destination saved chat_id={chat_id}")
        flash("Backup destination saved to DB. (You can also set via Telegram: /backup_here)", "success")
    else:
        log_event("ERROR", "web.backup_dest_set", f"Failed saving backup destination chat_id={chat_id}")
        flash("Failed to save destination. Check Logs.", "error")
    return redirect(url_for("backups_panel"))



@app.get("/cron/backup")
def cron_backup():
    """Cron-friendly endpoint to send backups to Telegram.

    Call: /cron/backup?key=CRON_SECRET
    """
    from ops_log import log_event

    key = (request.args.get("key") or "").strip()
    if CRON_SECRET and key != CRON_SECRET:
        return jsonify({"ok": False, "error": "unauthorized"}), 401

    chat_id = get_backup_chat_id(request.args.get("chat_id"))
    if not chat_id:
        return jsonify({"ok": False, "error": "BACKUP_TELEGRAM_CHAT_ID not set"}), 400

    artifacts, report = _make_backup_artifacts()

    mode = (request.args.get("mode") or "all").strip().lower()
    if mode in ("json","jsonzip","zip"):
        wanted = ["json_zip"]
    elif mode in ("excel","xlsx"):
        wanted = ["excel"]
    elif mode in ("logs","log"):
        wanted = ["logs_txt"]
    else:
        wanted = ["excel","json_zip","logs_txt"]

    sent = []
    errors = []

    caption_header = f"VO Tracker Backup • {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC • v{APP_VERSION}"
    for k in wanted:
        if k not in artifacts:
            continue
        fname, bts = artifacts[k]
        try:
            res = _tg_send_document(chat_id, fname, bts, caption=caption_header)
            sent.append({"kind": k, "file": fname, "message_id": res.get("message_id")})
        except Exception as e:
            errors.append({"kind": k, "file": fname, "error": str(e)})

    report["sent"] = sent
    report["send_errors"] = errors
    report["chat_id"] = chat_id

    level = "INFO" if not errors else "WARN"
    log_event(
        level,
        "cron.backup_send_tg",
        f"Cron TG backup chat={chat_id} sent={len(sent)} errors={len(errors)} token_tail={_tg_token_tail(BOT_TOKEN)}",
        traceback=json.dumps(report, ensure_ascii=False, indent=2),
    )

    return jsonify({"ok": len(errors) == 0, "sent": sent, "errors": errors, "generated_at": report.get("generated_at"), "version": APP_VERSION})


@app.get("/cron/admin_digest")
def cron_admin_digest():
    """Cron-friendly endpoint to send admin digest to Telegram."""
    key = (request.args.get("key") or "").strip()
    if CRON_SECRET and key != CRON_SECRET:
        return jsonify({"ok": False, "error": "unauthorized"}), 401

    enabled_db = (kv_get("admin_digest_enabled") or "").strip().lower()
    if enabled_db in {"0", "false", "no", "off"}:
        return jsonify({"ok": False, "error": "admin digest disabled"}), 400
    if enabled_db not in {"1", "true", "yes", "on"}:
        enabled_env = (os.getenv("ADMIN_DIGEST_ENABLED") or "1").strip().lower()
        if enabled_env in {"0", "false", "no", "off", ""}:
            return jsonify({"ok": False, "error": "admin digest disabled"}), 400

    if not BOT_ENABLED or not bot_app:
        return jsonify({"ok": False, "error": "bot disabled"}), 400

    chat_id = (request.args.get("chat_id") or "").strip()
    if not chat_id:
        chat_id = (os.getenv("ADMIN_DIGEST_CHAT_ID") or "").strip() or kv_get("digest_chat_id") or (os.getenv("ADMIN_TELEGRAM_CHAT_ID") or "").strip() or get_backup_chat_id()
    if not chat_id:
        return jsonify({"ok": False, "error": "digest destination not set"}), 400

    try:
        limit = int(request.args.get("limit", "5"))
    except Exception:
        limit = 5
    limit = max(1, min(limit, 10))

    text_msg = build_admin_digest_text(priority_limit=limit)
    try:
        run_async(bot_app.bot.send_message(chat_id=int(chat_id), text=text_msg, disable_web_page_preview=True))
        log_event("INFO", "cron.admin_digest", f"Cron admin digest sent chat={chat_id} limit={limit}")
        return jsonify({"ok": True, "chat_id": chat_id, "limit": limit, "version": APP_VERSION})
    except Exception as e:
        log_event("ERROR", "cron.admin_digest", f"Cron admin digest failed chat={chat_id}: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500
@app.get("/backups")
@login_required
@require_admin
def backups_panel():
    """Backup Center: quick access to exports + recent status."""
    try:
        rows = (
            db.session.execute(
                sql_text(
                    """
                    SELECT ts, level, source, message
                    FROM system_logs
                    WHERE source IN ('web.export_excel','web.backup_json','web.backup_send_tg','cron.backup_send_tg')
                    ORDER BY id DESC
                    LIMIT 20
                    """
                )
            )
            .mappings()
            .all()
        )
    except Exception:
        try:
            db.session.rollback()
        except Exception:
            pass
        rows = []

    entries = []
    for r in rows:
        ts = r.get("ts")
        entries.append(
            {
                "ts": (ts.strftime("%Y-%m-%d %H:%M:%S") if hasattr(ts, "strftime") else (str(ts) if ts else "")),
                "level": r.get("level") or "INFO",
                "source": r.get("source") or "",
                "message": r.get("message") or "",
            }
        )

    return render_template(
        "backups.html",
        title="Backups",
        entries=entries,
        backup_chat_env=bool(BACKUP_TELEGRAM_CHAT_ID),
        backup_chat_db=kv_get("backup_chat_id"),
        backup_chat_effective=get_backup_chat_id(),
        backup_chat_configured=bool(get_backup_chat_id()),
    )


@app.get("/tips")
@login_required
@require_admin
def tips_panel():
    """Short workflow tips (web)."""
    return render_template("tips.html", title="Tips")


@app.get("/export/logs.txt")
@login_required
@require_admin
def export_logs_txt():
    """Download system logs as a text file (human-readable)."""
    try:
        limit = int(request.args.get("limit") or 2000)
    except Exception:
        limit = 2000
    limit = max(1, min(limit, int(os.getenv("EXPORT_MAX_LOGS", "5000"))))

    try:
        items = fetch_logs(limit=limit)
    except Exception as e:
        try:
            db.session.rollback()
        except Exception:
            pass
        items = [{"ts": "", "level": "ERROR", "source": "web.export_logs", "message": f"Failed: {e}", "traceback": ""}]

    out_lines = []
    for it in items:
        out_lines.append(f"[{it.get('level','INFO')}] {it.get('ts','')} ({it.get('source','')})")
        out_lines.append(it.get("message") or "")
        tb = (it.get("traceback") or "").strip()
        if tb:
            out_lines.append(tb)
        out_lines.append("" + "-" * 60)

    content = "\n".join(out_lines)
    fname = f"system_logs_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}Z.txt"
    return Response(
        content,
        mimetype="text/plain; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# --------------------------------------------------
# RESTORE (from JSON ZIP backup)
# --------------------------------------------------
@app.get("/restore")
@login_required
@require_admin
def restore_panel():
    """Restore Center: upload a JSON ZIP backup and restore DB content."""
    # Recent restore logs
    try:
        rows = (
            db.session.execute(
                sql_text(
                    """
                    SELECT ts, level, source, message
                    FROM system_logs
                    WHERE source IN ('web.restore_dry_run','web.restore_run')
                    ORDER BY id DESC
                    LIMIT 20
                    """
                )
            )
            .mappings()
            .all()
        )
    except Exception:
        try:
            db.session.rollback()
        except Exception:
            pass
        rows = []

    entries = []
    for r in rows:
        ts = r.get("ts")
        entries.append({
            "ts": (ts.strftime("%Y-%m-%d %H:%M:%S") if hasattr(ts, "strftime") else (str(ts) if ts else "")),
            "level": r.get("level") or "INFO",
            "source": r.get("source") or "",
            "message": r.get("message") or "",
        })

    try:
        return render_template("restore.html", entries=entries, dry_report=None, run_report=None)
    except TemplateNotFound:
        # Fallback: minimal HTML so app does not crash if templates were not deployed
        return "<h1>Restore template missing</h1><p>Deploy templates/restore.html</p>", 500


@app.post("/restore/dry_run")
@login_required
@require_admin
def restore_dry_run():
    f = request.files.get("backup_zip")
    if not f:
        flash("❌ Please upload a JSON ZIP backup")
        return redirect(url_for("restore_panel"))

    include_admin = bool(request.form.get("include_admin"))
    include_logs = bool(request.form.get("include_logs"))
    zip_bytes = f.read() or b""
    if not zip_bytes:
        flash("❌ Empty file")
        return redirect(url_for("restore_panel"))

    try:
        rep = restore_dry_run_dynamic(db.engine, zip_bytes)
        # Apply the same inclusion filters as restore
        # (Only affects display; restore will enforce again.)
        if not include_admin:
            rep["table_reports"] = [r for r in rep.get("table_reports", []) if r.get("table") not in ("admin_user", "admin_telegram_user")]
        if not include_logs:
            rep["table_reports"] = [r for r in rep.get("table_reports", []) if r.get("table") != "system_logs"]

        msg = f"Dry run OK. tables={rep.get('tables_in_backup')} (admin={include_admin}, logs={include_logs})"
        log_event("INFO", "web.restore_dry_run", msg, traceback=json.dumps(rep, ensure_ascii=False, indent=2))
        flash("✅ Dry run completed. See results below.")
        # show recent restore logs too
        return render_template("restore.html", entries=[], dry_report=rep, run_report=None)
    except TemplateNotFound:
        return "<h1>Restore template missing</h1><p>Deploy templates/restore.html</p>", 500
    except Exception as e:
        try:
            db.session.rollback()
        except Exception:
            pass
        log_event("ERROR", "web.restore_dry_run", f"Dry run FAILED: {e}")
        flash(f"❌ Dry run failed: {e}")
        return redirect(url_for("restore_panel"))


@app.post("/restore/run")
@login_required
@require_admin
def restore_run():
    f = request.files.get("backup_zip")
    if not f:
        flash("❌ Please upload a JSON ZIP backup")
        return redirect(url_for("restore_panel"))

    confirm = (request.form.get("confirm") or "").strip().upper()
    if confirm != "RESTORE":
        flash("❌ Confirmation text must be RESTORE")
        return redirect(url_for("restore_panel"))

    include_admin = bool(request.form.get("include_admin"))
    include_logs = bool(request.form.get("include_logs"))
    mode = (request.form.get("mode") or "replace").strip().lower()
    if mode not in ("replace", "append", "merge"):
        mode = "replace"
    # Optional: restore only selected tables
    tables_selected = request.form.getlist("tables")
    tables_csv = (request.form.get("tables_csv") or "").strip()
    if (not tables_selected) and tables_csv:
        tables_selected = [t.strip() for t in tables_csv.split(",") if t.strip()]
    only_tables = tables_selected or None
    zip_bytes = f.read() or b""
    if not zip_bytes:
        flash("❌ Empty file")
        return redirect(url_for("restore_panel"))

    try:
        rep = restore_from_backup_zip(
            db.engine,
            zip_bytes,
            include_admin=include_admin,
            include_logs=include_logs,
            mode=mode,
            only_tables=only_tables,
        )
        level = "INFO" if not rep.get("errors") else "WARN"
        msg = f"Restore finished. tables={len(rep.get('tables', []))} errors={len(rep.get('errors', []))}"
        log_event(level, "web.restore_run", msg, traceback=json.dumps(rep, ensure_ascii=False, indent=2))
        flash("✅ Restore completed. Review the report below.")
        return render_template("restore.html", entries=[], dry_report=None, run_report=rep)
    except TemplateNotFound:
        return "<h1>Restore template missing</h1><p>Deploy templates/restore.html</p>", 500
    except Exception as e:
        try:
            db.session.rollback()
        except Exception:
            pass
        log_event("ERROR", "web.restore_run", f"Restore FAILED: {e}")
        flash(f"❌ Restore failed: {e}")
        return redirect(url_for("restore_panel"))



@app.route("/assignments/rerun", methods=["POST"])
@login_required
@require_admin
def assignments_rerun():
    project = request.form.get("project")
    mode = normalize_priority_mode(request.form.get("mode", "urgent"))
    urgent = priority_mode_urgent_only(mode)
    default_deadline = priority_mode_deadline(mode)
    text = request.form.get("bulk_text")

    parsed = parse_lines(text)
    if not parsed:
        flash("❌ Nothing parsed")
        return redirect(url_for("assignments_view", project=project))

    # One role -> one VO (aggregate duplicates like man-1 repeated per character).
    agg: dict[str, int] = {}
    for role, lines in parsed:
        agg[role] = agg.get(role, 0) + int(lines or 0)
    parsed = sorted(agg.items(), key=lambda x: x[0])

    movie = Movie.query.filter_by(code=(project or "").strip().upper()).first()

    Assignment.query.filter_by(project=project).delete()
    db.session.commit()

    # Workload across projects (used by pick_vo preference rules).
    from sqlalchemy import func
    rows = (
        db.session.query(Assignment.vo, func.count(func.distinct(Assignment.project)))
        .group_by(Assignment.vo)
        .all()
    )
    project_counts = {vo: int(cnt or 0) for vo, cnt in rows}

    load = movie_load(project)
    used = set()

    for role, lines in parsed:
        gender = role_gender(role)
        q = VOTeam.query.filter_by(active=True, gender=gender)
        if urgent:
            q = q.filter_by(urgent_ok=True)

        picked = pick_vo(q.all(), used, load, project_counts)
        if not picked:
            continue

        used.add(picked.name)
        db.session.add(
            Assignment(
                project=project,
                movie_id=(movie.id if movie else None),
                vo=picked.name,
                role=role,
                lines=lines,
                urgent=bool(urgent),
                priority_mode=mode,
                deadline_at=default_deadline,
            )
        )
        db.session.commit()

    if movie:
        record_movie_event(movie, "RERUN_ASSIGN", f"Web re-ran auto assign", detail=f"mode={mode} • urgent={urgent} • roles={len(parsed)}", actor_source="web", actor_name="assignments_rerun")
    db.session.commit()
    flash("✅ Re-run completed")
    return redirect(url_for("assignments_view", project=project))

# --------------------------------------------------

# --------------------------------------------------
# VO TEAM (fix /vo-team 404)
# --------------------------------------------------
@app.route("/vo-team", methods=["GET", "POST"])
@login_required
@require_admin
def vo_team_panel():
    if request.method == "POST":
        action = (request.form.get("action") or "").strip()

        if action == "add":
            name = (request.form.get("name") or "").strip()
            tg_username = (request.form.get("tg_username") or "").strip().lstrip("@").strip() or None
            tg_user_id_raw = (request.form.get("tg_user_id") or "").strip()
            tg_user_id = int(tg_user_id_raw) if tg_user_id_raw.isdigit() else None
            gender = (request.form.get("gender") or "male").strip().lower()
            level = (request.form.get("level") or "trained_new").strip()
            speed = (request.form.get("speed") or "normal").strip()
            urgent_ok = bool(request.form.get("urgent_ok"))
            active = bool(request.form.get("active"))

            if not name:
                flash("❌ Name required")
                return redirect(url_for("vo_team_panel"))

            existing = VOTeam.query.filter_by(name=name).first()
            if existing:
                flash("⚠️ Name already exists")
                return redirect(url_for("vo_team_panel"))

            db.session.add(VOTeam(
                name=name,
                tg_username=tg_username,
                tg_user_id=tg_user_id,
                gender=gender,
                level=level,
                speed=speed,
                urgent_ok=urgent_ok,
                active=active,
            ))
            db.session.commit()
            flash("✅ Added")
            return redirect(url_for("vo_team_panel"))

        if action == "update":
            vid = request.form.get("id")
            row = db.session.get(VOTeam, int(vid)) if vid and vid.isdigit() else None
            if not row:
                flash("❌ Not found")
                return redirect(url_for("vo_team_panel"))

            row.gender = (request.form.get("gender") or row.gender).strip().lower()
            row.level = (request.form.get("level") or row.level).strip()
            row.speed = (request.form.get("speed") or row.speed).strip()
            row.tg_username = (request.form.get("tg_username") or "").strip().lstrip("@").strip() or None
            tg_user_id_raw = (request.form.get("tg_user_id") or "").strip()
            if tg_user_id_raw.isdigit():
                row.tg_user_id = int(tg_user_id_raw)
            row.urgent_ok = bool(request.form.get("urgent_ok"))
            row.active = bool(request.form.get("active"))
            db.session.commit()
            flash("✅ Updated")
            return redirect(url_for("vo_team_panel"))

        if action == "delete":
            vid = request.form.get("id")
            row = db.session.get(VOTeam, int(vid)) if vid and vid.isdigit() else None
            if not row:
                flash("❌ Not found")
                return redirect(url_for("vo_team_panel"))

            # Safety: only allow deleting VO who is already inactive
            if row.active:
                flash("⚠️ Set Active off first before deleting")
                return redirect(url_for("vo_team_panel"))

            db.session.delete(row)
            db.session.commit()
            flash("🗑️ Deleted")
            return redirect(url_for("vo_team_panel"))

        flash("❌ Unknown action")
        return redirect(url_for("vo_team_panel"))

    members = VOTeam.query.order_by(VOTeam.gender.asc(), VOTeam.level.asc(), VOTeam.name.asc()).all()
    return render_template("vo_team.html", members=members)


# --------------------------------------------------
# VO WORKLOAD (deadline-only overdue)
# --------------------------------------------------


@app.get("/vo/<int:vo_id>")
@login_required
@require_admin
def vo_workload(vo_id: int):
    view = (request.args.get("view") or "all").strip().lower()
    if view not in {"all", "pending", "overdue"}:
        view = "all"

    vo = db.session.get(VOTeam, int(vo_id))
    if not vo:
        flash("❌ VO not found")
        return redirect(url_for("dashboard"))

    now_utc = datetime.utcnow()

    items = Assignment.query.filter_by(vo=vo.name).order_by(Assignment.project.desc(), Assignment.role.asc()).all()

    # done sets
    try:
        subs = VORoleSubmission.query.with_entities(VORoleSubmission.movie, VORoleSubmission.vo, VORoleSubmission.role).all()
    except Exception:
        subs = []
    done_triplet = {(m, v, r) for (m, v, r) in subs}
    done_pair = {(m, r) for (m, v, r) in subs}

    rows = []
    total = len(items)
    done = 0
    pending = 0
    late = 0
    late_lines = 0

    for a in items:
        is_done = ((a.project, a.vo, a.role) in done_triplet) or ((a.project, a.role) in done_pair)
        status = "DONE" if is_done else "PENDING"
        if not is_done:
            pending += 1
            if a.deadline_at and a.deadline_at < now_utc:
                status = "OVERDUE"
                late += 1
                late_lines += int(a.lines or 0)
        else:
            done += 1

        # filter
        if view == "pending" and status != "PENDING":
            continue
        if view == "overdue" and status != "OVERDUE":
            continue

        deadline_input = ""
        if a.deadline_at:
            deadline_input = utc_to_myt(a.deadline_at).strftime("%Y-%m-%dT%H:%M")

        rows.append(
            {
                "id": a.id,
                "project": a.project,
                "role": a.role,
                "lines": int(a.lines or 0),
                "status": status,
                "deadline_myt": fmt_myt(a.deadline_at),
                "deadline_input": deadline_input,
            }
        )

    summary = {"total": total, "pending": pending, "late": late, "done": done, "late_lines": late_lines}
    return render_template("vo_workload.html", vo=vo, view=view, rows=rows, summary=summary)


@app.post("/assignments/<int:assignment_id>/deadline")
@login_required
@require_admin
def assignment_set_deadline(assignment_id: int):
    a = db.session.get(Assignment, int(assignment_id))
    if not a:
        flash("❌ Assignment not found")
        return redirect(url_for("dashboard"))

    dt_utc = parse_myt_datetime_local(request.form.get("deadline"))
    a.deadline_at = dt_utc
    db.session.commit()
    flash("✅ Deadline saved")
    nxt = (request.form.get("next") or "").strip()
    if nxt:
        return redirect(nxt)
    vo = VOTeam.query.filter_by(name=a.vo).first()
    return redirect(url_for("vo_workload", vo_id=vo.id) if vo else url_for("dashboard"))


@app.post("/assignments/<int:assignment_id>/remind")
@login_required
@require_admin
def assignment_remind(assignment_id: int):
    a = db.session.get(Assignment, int(assignment_id))
    if not a:
        flash("❌ Assignment not found")
        return redirect(url_for("dashboard"))

    if not BOT_ENABLED or not bot_app:
        flash("⚠️ Telegram bot disabled (missing BOT_TOKEN)")
        return redirect(request.form.get("next") or url_for("dashboard"))

    vo = VOTeam.query.filter_by(name=a.vo).first()
    if not vo or not vo.tg_user_id:
        flash("⚠️ Cannot DM this VO: missing tg_user_id in VO Team roster")
        # also ping admin chat if configured
        try:
            if ADMIN_TELEGRAM_CHAT_ID:
                run_async(bot_app.bot.send_message(chat_id=int(ADMIN_TELEGRAM_CHAT_ID), text=f"⚠️ Remind failed: VO '{a.vo}' has no tg_user_id (assignment {a.project} {a.role})."))
        except Exception:
            pass
        return redirect(request.form.get("next") or url_for("dashboard"))

    deadline_txt = fmt_myt(a.deadline_at)
    msg = "\n".join(
        [
            "⏰ *Reminder — VO overdue*",
            f"Movie: `{a.project}`",
            f"Role: `{a.role}` ({int(a.lines or 0)} lines)",
            f"Deadline: *{deadline_txt}*" if a.deadline_at else "Deadline: -",
            "Please submit ASAP. If you need more time, tell admin.",
        ]
    )

    try:
        run_async(bot_app.bot.send_message(chat_id=int(vo.tg_user_id), text=msg, parse_mode="Markdown"))
        a.last_reminded_at = datetime.utcnow()
        a.remind_count = int(a.remind_count or 0) + 1
        db.session.commit()
        log_event("INFO", "web.remind", f"VO remind sent to={vo.name} movie={a.project} role={a.role}")
        flash("✅ Reminder sent")
    except Exception as e:
        db.session.rollback()
        flash(f"❌ Reminder failed: {e}")

    return redirect(request.form.get("next") or url_for("dashboard"))


# --------------------------------------------------
# TRANSLATORS ROSTER
# --------------------------------------------------
@app.route("/translators", methods=["GET", "POST"])
@login_required
@require_admin
def translators_panel():
    if request.method == "POST":
        action = (request.form.get("action") or "").strip()

        if action == "add":
            name = (request.form.get("name") or "").strip()
            tg_username = (request.form.get("tg_username") or "").strip().lstrip("@").strip() or None
            tg_user_id_raw = (request.form.get("tg_user_id") or "").strip()
            tg_user_id = int(tg_user_id_raw) if tg_user_id_raw.isdigit() else None
            languages = (request.form.get("languages") or "").strip() or None
            active = bool(request.form.get("active"))

            if not name:
                flash("❌ Name required")
                return redirect(url_for("translators_panel"))

            existing = Translator.query.filter_by(name=name).first()
            if existing:
                flash("⚠️ Name already exists")
                return redirect(url_for("translators_panel"))

            row = Translator(
                name=name,
                tg_username=tg_username,
                tg_user_id=tg_user_id,
                languages=languages,
                active=active,
            )
            db.session.add(row)
            db.session.commit()
            log_event("INFO", "web.translators", f"ADD name={name} username={tg_username or '-'} id={tg_user_id or '-'}")
            flash("✅ Added")
            return redirect(url_for("translators_panel"))

        if action == "update":
            tid = (request.form.get("id") or "").strip()
            row = db.session.get(Translator, int(tid)) if tid.isdigit() else None
            if not row:
                flash("❌ Not found")
                return redirect(url_for("translators_panel"))

            tg_username = (request.form.get("tg_username") or "").strip().lstrip("@").strip() or None
            tg_user_id_raw = (request.form.get("tg_user_id") or "").strip()
            tg_user_id = int(tg_user_id_raw) if tg_user_id_raw.isdigit() else None
            languages = (request.form.get("languages") or "").strip() or None

            row.tg_username = tg_username
            if tg_user_id:
                row.tg_user_id = tg_user_id
            row.languages = languages
            row.active = bool(request.form.get("active"))
            db.session.commit()
            log_event("INFO", "web.translators", f"UPDATE name={row.name}")
            flash("✅ Saved")
            return redirect(url_for("translators_panel"))

        if action == "delete":
            tid = (request.form.get("id") or "").strip()
            row = db.session.get(Translator, int(tid)) if tid.isdigit() else None
            if not row:
                flash("❌ Not found")
                return redirect(url_for("translators_panel"))

            if row.active:
                flash("⚠️ Set Active off first before deleting")
                return redirect(url_for("translators_panel"))

            name = row.name
            db.session.delete(row)
            db.session.commit()
            log_event("WARN", "web.translators", f"DELETE name={name}")
            flash("🗑️ Deleted")
            return redirect(url_for("translators_panel"))

        flash("❌ Unknown action")
        return redirect(url_for("translators_panel"))

    translators = Translator.query.order_by(Translator.name.asc()).all()

    now_utc = datetime.utcnow()
    # Stats: Total/Pending/Late from TranslationTask (fallback to zero if none)
    try:
        all_tasks = TranslationTask.query.all()
    except Exception:
        all_tasks = []

    stats = {}
    for t in translators:
        rel = [x for x in all_tasks if x.translator_id == t.id]
        if not rel and (t.name or t.tg_username):
            keyset = set([ (t.name or "").strip().lower(), (t.tg_username or "").strip().lower() ])
            rel = [x for x in all_tasks if (x.translator_name or "").strip().lower() in keyset]

        total = len(rel)
        pending = sum(1 for x in rel if (x.status or "").upper() != "COMPLETED")
        late = sum(1 for x in rel if (x.status or "").upper() != "COMPLETED" and x.deadline_at and x.deadline_at < now_utc)
        stats[t.id] = {"total": int(total), "pending": int(pending), "late": int(late)}

    return render_template("translators.html", translators=translators, stats=stats)


# --------------------------------------------------
# TRANSLATOR WORKLOAD (TranslationTask)
# --------------------------------------------------


@app.get("/translator/<int:translator_id>")
@login_required
@require_admin
def translator_workload(translator_id: int):
    view = (request.args.get("view") or "all").strip().lower()
    if view not in {"all", "pending", "overdue"}:
        view = "all"

    tr = db.session.get(Translator, int(translator_id))
    if not tr:
        flash("❌ Translator not found")
        return redirect(url_for("translators_panel"))

    now_utc = datetime.utcnow()

    tasks = TranslationTask.query.filter_by(translator_id=tr.id).order_by(TranslationTask.id.desc()).all()
    # fallback: if tasks are legacy-only by name
    if not tasks:
        keyset = set([ (tr.name or "").strip().lower(), (tr.tg_username or "").strip().lower() ])
        tasks = [x for x in TranslationTask.query.all() if (x.translator_name or "").strip().lower() in keyset]

    rows = []
    total = len(tasks)
    done = 0
    pending = 0
    late = 0

    for t in tasks:
        st = (t.status or "SENT").upper()
        is_done = st == "COMPLETED"
        is_overdue = (not is_done) and t.deadline_at and t.deadline_at < now_utc

        status_label = "COMPLETED" if is_done else ("OVERDUE" if is_overdue else st)
        if is_done:
            done += 1
        else:
            pending += 1
            if is_overdue:
                late += 1

        if view == "pending" and (is_done or is_overdue):
            continue
        if view == "overdue" and not is_overdue:
            continue

        deadline_input = ""
        if t.deadline_at:
            deadline_input = utc_to_myt(t.deadline_at).strftime("%Y-%m-%dT%H:%M")

        rows.append(
            {
                "id": t.id,
                "movie_code": t.movie_code,
                "title": t.title,
                "year": t.year,
                "lang": t.lang,
                "status": status_label,
                "deadline_myt": fmt_myt(t.deadline_at),
                "deadline_input": deadline_input,
                "sent_utc": t.sent_at.strftime("%Y-%m-%d %H:%M") + " UTC" if t.sent_at else "-",
                "completed_utc": t.completed_at.strftime("%Y-%m-%d %H:%M") + " UTC" if t.completed_at else "-",
            }
        )

    summary = {"total": total, "pending": pending, "late": late, "done": done}
    return render_template("translator_workload.html", tr=tr, view=view, rows=rows, summary=summary)


@app.post("/translation_task/<int:task_id>/deadline")
@login_required
@require_admin
def translation_task_set_deadline(task_id: int):
    t = db.session.get(TranslationTask, int(task_id))
    if not t:
        flash("❌ Task not found")
        return redirect(url_for("translators_panel"))
    t.deadline_at = parse_myt_datetime_local(request.form.get("deadline"))
    db.session.commit()
    flash("✅ Deadline saved")
    nxt = (request.form.get("next") or "").strip()
    if nxt:
        return redirect(nxt)
    if t.translator_id:
        return redirect(url_for("translator_workload", translator_id=t.translator_id))
    return redirect(url_for("translators_panel"))


@app.post("/translation_task/<int:task_id>/remind")
@login_required
@require_admin
def translation_task_remind(task_id: int):
    t = db.session.get(TranslationTask, int(task_id))
    if not t:
        flash("❌ Task not found")
        return redirect(url_for("translators_panel"))

    if not BOT_ENABLED or not bot_app:
        flash("⚠️ Telegram bot disabled (missing BOT_TOKEN)")
        return redirect(request.form.get("next") or url_for("translators_panel"))

    tr = db.session.get(Translator, int(t.translator_id)) if t.translator_id else None
    if not tr or not tr.tg_user_id:
        flash("⚠️ Cannot DM this translator: missing tg_user_id in Translators roster")
        try:
            if ADMIN_TELEGRAM_CHAT_ID:
                run_async(bot_app.bot.send_message(chat_id=int(ADMIN_TELEGRAM_CHAT_ID), text=f"⚠️ Remind failed: translator task {t.id} has no tg_user_id."))
        except Exception:
            pass
        return redirect(request.form.get("next") or url_for("translators_panel"))

    deadline_txt = fmt_myt(t.deadline_at)
    msg = "\n".join(
        [
            "⏰ *Reminder — Translation overdue*",
            f"Movie: `{t.movie_code or '-'}`",
            f"Title: *{t.title or '-'}*" + (f" ({t.year})" if t.year else ""),
            f"Lang: *{(t.lang or '').upper() or '-'}*",
            f"Deadline: *{deadline_txt}*" if t.deadline_at else "Deadline: -",
            "Please submit the .srt ASAP. If you need more time, tell admin.",
        ]
    )

    try:
        run_async(bot_app.bot.send_message(chat_id=int(tr.tg_user_id), text=msg, parse_mode="Markdown"))
        t.last_reminded_at = datetime.utcnow()
        t.remind_count = int(t.remind_count or 0) + 1
        db.session.commit()
        log_event("INFO", "web.remind", f"TR remind sent to={tr.name} movie={t.movie_code} task={t.id}")
        flash("✅ Reminder sent")
    except Exception as e:
        db.session.rollback()
        flash(f"❌ Reminder failed: {e}")

    return redirect(request.form.get("next") or url_for("translator_workload", translator_id=tr.id))

# QUEUE PANEL
# --------------------------------------------------
def notify_submitter(sub: TranslationSubmission, text: str):
    if not sub.submitter_id:
        return
    if not BOT_ENABLED or not bot_app:
        return
    try:
        run_async(
            bot_app.bot.send_message(
                chat_id=int(sub.submitter_id),
                text=text,
            )
        )
    except Exception as e:
        log.warning("Notify failed sub=%s user=%s err=%s", sub.id, sub.submitter_id, e)


@app.route("/queue")
@login_required
@require_admin
def queue_panel():
    rows = TranslationSubmission.query.order_by(TranslationSubmission.submitted_at.desc()).limit(200).all()
    return render_template("queue.html", rows=rows)


# ✅ FIXED: lock + reject note required + Movie sync
@app.route("/queue/<int:sid>/<action>", methods=["POST"])
@login_required
@require_admin
def queue_action(sid, action):
    sub = db.session.get(TranslationSubmission, sid)
    if not sub:
        return ("Not found", 404)

    # 🔒 lock: once DONE/REJECTED, cannot change
    if sub.status in ("DONE", "REJECTED"):
        flash(f"🔒 Locked: submission #{sub.id} already {sub.status}")
        return redirect(url_for("queue_panel"))

    movie = db.session.get(Movie, sub.movie_id) if sub.movie_id else None

    def disp_movie() -> str:
        if movie:
            return f"{movie.title} ({movie.year})"
        # fallback
        return sub.movie

    if action == "in_qa":
        sub.status = "IN_QA"
        if movie:
            movie.status = "IN_QA"
            movie.updated_at = datetime.utcnow()
        msg = f"🧪 QA started for {disp_movie()}"

    elif action == "done":
        sub.status = "DONE"
        if movie:
            movie.status = "SUBMITTED"
            movie.submitted_at = movie.submitted_at or datetime.utcnow()
            movie.updated_at = datetime.utcnow()
        msg = f"✅ Approved: {disp_movie()}"

    elif action == "reject":
        note = (sub.note or "").strip()
        if not note:
            flash("❌ REJECT requires a note. Save note first.")
            return redirect(url_for("queue_panel"))

        sub.status = "REJECTED"
        if movie:
            movie.status = "NEED_REWORK"
            movie.updated_at = datetime.utcnow()
        # Show full movie name instead of code when possible
        disp = sub.movie
        try:
            m2 = None
            if getattr(sub, "movie_id", None):
                m2 = Movie.query.get(int(sub.movie_id))
            if not m2 and sub.movie:
                m2 = Movie.query.filter_by(code=sub.movie).first()
            if m2:
                disp = f"{m2.title} ({m2.year})"
        except Exception:
            pass
        msg = f"❌ Rejected: {disp}\n\nNote: {note}"

    else:
        return ("Bad action", 400)

    db.session.commit()
    notify_submitter(sub, msg)
    flash(f"Updated #{sub.id} → {sub.status}")
    return redirect(url_for("queue_panel"))


@app.route("/queue/<int:sid>/note", methods=["POST"])
@login_required
@require_admin
def queue_note(sid):
    sub = db.session.get(TranslationSubmission, sid)
    if not sub:
        return ("Not found", 404)
    sub.note = request.form.get("note", "")
    db.session.commit()
    flash("Note saved")
    return redirect(url_for("queue_panel"))


@app.route("/queue/<int:sid>/delete", methods=["POST"])
@login_required
@require_admin
def queue_delete(sid):
    sub = db.session.get(TranslationSubmission, sid)
    if not sub:
        return ("Not found", 404)
    db.session.delete(sub)
    db.session.commit()
    flash(f"🗑 Deleted submission #{sid}")
    return redirect(url_for("queue_panel"))


@app.route("/queue/reset", methods=["POST"])
@login_required
@require_admin
def queue_reset():
    mode = (request.form.get("mode") or "").strip()
    movie_filter = (request.form.get("movie") or "").strip()
    confirm = bool(request.form.get("confirm"))

    if not confirm:
        flash("❌ Tick CONFIRM first")
        return redirect(url_for("queue_panel"))

    q = TranslationSubmission.query
    if movie_filter:
        q = q.filter(TranslationSubmission.movie.ilike(f"%{movie_filter}%"))

    affected = 0

    if mode == "unlock":
        rows = q.filter(TranslationSubmission.status.in_(["DONE", "REJECTED", "IN_QA"])).all()
        for sub in rows:
            sub.status = "READY_FOR_QA"
            # keep note, but unlock actions
            if sub.movie_id:
                m = db.session.get(Movie, sub.movie_id)
                if m:
                    m.status = "READY_FOR_QA"
                    m.updated_at = datetime.utcnow()
            affected += 1
        db.session.commit()
        log_event("WARN", "web.queue", f"RESET unlock count={affected} filter={movie_filter or '-'}")
        flash(f"✅ Unlocked {affected} item(s) → READY_FOR_QA")

    elif mode == "purge_done_rejected":
        affected = q.filter(TranslationSubmission.status.in_(["DONE", "REJECTED"])).delete(synchronize_session=False)
        db.session.commit()
        log_event("WARN", "web.queue", f"RESET purge_done_rejected count={affected} filter={movie_filter or '-'}")
        flash(f"🗑️ Deleted {affected} DONE/REJECTED item(s)")

    elif mode == "purge_all":
        affected = q.delete(synchronize_session=False)
        db.session.commit()
        log_event("ERROR", "web.queue", f"RESET purge_all count={affected} filter={movie_filter or '-'}")
        flash(f"🗑️ Deleted {affected} item(s)")

    else:
        flash("❌ Unknown reset mode")

    return redirect(url_for("queue_panel"))

# --------------------------------------------------
# TELEGRAM PANEL
# --------------------------------------------------
@app.route("/telegram", methods=["GET", "POST"])
@login_required
@require_admin
def telegram_panel():
    webhook_url = f"{RENDER_EXTERNAL_URL}/webhook/{WEBHOOK_SECRET}" if RENDER_EXTERNAL_URL else ""

    submitters = (
        db.session.query(TranslationSubmission.submitter_id, TranslationSubmission.submitter_username)
        .filter(TranslationSubmission.submitter_id.isnot(None))
        .group_by(TranslationSubmission.submitter_id, TranslationSubmission.submitter_username)
        .all()
    )

    if request.method == "POST":
        action = request.form.get("action")

        if action == "set_webhook":
            if not BOT_ENABLED or not bot_app:
                flash("❌ Bot disabled (missing BOT_TOKEN)")
                return redirect(url_for("telegram_panel"))
            if not webhook_url:
                flash("❌ Missing RENDER_EXTERNAL_URL")
                return redirect(url_for("telegram_panel"))
            run_async(bot_app.bot.set_webhook(webhook_url))
            flash("Webhook set")
            return redirect(url_for("telegram_panel"))

        if action == "send":
            if not BOT_ENABLED or not bot_app:
                flash("❌ Bot disabled (missing BOT_TOKEN)")
                return redirect(url_for("telegram_panel"))

            target = request.form.get("target_mode")
            text_msg = request.form.get("message", "").strip()

            targets = []
            if target == "drop" and DROP_CHAT_ID:
                targets = [int(DROP_CHAT_ID)]
            elif target == "admin" and ADMIN_TELEGRAM_CHAT_ID:
                targets = [int(ADMIN_TELEGRAM_CHAT_ID)]
            elif target == "submitters":
                targets = [int(sid) for sid, _ in submitters if sid]

            sent = 0
            for cid in targets:
                try:
                    run_async(bot_app.bot.send_message(chat_id=cid, text=text_msg))
                    sent += 1
                except Exception:
                    pass

            flash(f"Sent to {sent}/{len(targets)} chat(s)")
            return redirect(url_for("telegram_panel"))

    return render_template(
        "telegram.html",
        webhook_url=webhook_url,
        submitters=submitters,
        drop_chat_id=DROP_CHAT_ID,
        admin_chat_id=ADMIN_TELEGRAM_CHAT_ID,
    )

# --------------------------------------------------
# TELEGRAM WEBHOOK
# --------------------------------------------------
async def _handle_update(update: Update):
    if not BOT_ENABLED or not bot_app:
        return
    with app.app_context():
        await bot_app.process_update(update)


@app.route(f"/webhook/{WEBHOOK_SECRET}", methods=["POST"])
def telegram_webhook():
    if not BOT_ENABLED or not bot_app:
        return "OK", 200
    try:
        payload = request.get_json(force=True)
        update = Update.de_json(payload, bot_app.bot)
        run_bot_coro(_handle_update(update))
    except Exception as e:
        log.exception("Webhook failed: %s", e)
    return "OK", 200


@app.route("/setup-webhook")
@login_required
@require_admin
def setup_webhook():
    if not BOT_ENABLED or not bot_app:
        return "Bot disabled (missing BOT_TOKEN)", 400
    if not RENDER_EXTERNAL_URL:
        return "Missing RENDER_EXTERNAL_URL", 400
    webhook_url = f"{RENDER_EXTERNAL_URL}/webhook/{WEBHOOK_SECRET}"
    try:
        run_async(bot_app.bot.set_webhook(webhook_url))
        return f"Webhook set to {webhook_url}", 200
    except Exception as e:
        log.exception("set_webhook failed: %s", e)
        return "Failed", 500


# --------------------------------------------------
# Local dev entrypoint
# --------------------------------------------------
if __name__ == "__main__":
    # Render uses gunicorn (app:app). This block is only for local runs.
    host = os.getenv("HOST", "0.0.0.0")
    port = int(os.getenv("PORT", "5000"))
    app.run(host=host, port=port, debug=True)
