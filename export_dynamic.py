"""Robust export/backup utilities.

These functions intentionally do NOT rely on SQLAlchemy ORM models for reading,
because a common failure mode in free-tier deployments is partial schema drift
(models define columns that do not exist yet).

We instead introspect the live database schema and export whatever exists.
"""

from __future__ import annotations

import base64
import json
import os
import zipfile
from datetime import date, datetime, time, timezone
from decimal import Decimal
from io import BytesIO
from typing import Any, Dict, List, Tuple

from sqlalchemy import inspect, text as sql_text
from sqlalchemy.engine import Engine

from openpyxl import Workbook


def _iso(v: Any) -> Any:
    if v is None:
        return None
    if isinstance(v, datetime):
        try:
            if v.tzinfo is None:
                v = v.replace(tzinfo=timezone.utc)
            return v.isoformat()
        except Exception:
            return str(v)
    if isinstance(v, (date, time)):
        try:
            return v.isoformat()
        except Exception:
            return str(v)
    if isinstance(v, bytes):
        return {"__bytes__": base64.b64encode(v).decode("ascii")}
    if isinstance(v, Decimal):
        # Keep exactness by exporting string.
        return str(v)
    # JSON-able types
    if isinstance(v, (int, float, bool, str, list, dict)):
        return v
    return str(v)


def _safe_cell(v: Any) -> Any:
    """Convert values to Excel-friendly cells."""
    v = _iso(v)
    if isinstance(v, dict) or isinstance(v, list):
        return json.dumps(v, ensure_ascii=False)
    return v


def list_tables(engine: Engine) -> List[str]:
    insp = inspect(engine)
    return sorted(insp.get_table_names())


def fetch_table(engine: Engine, table: str, max_rows: int | None = None) -> Tuple[List[str], List[Dict[str, Any]]]:
    """Return (columns, rows) for a table using raw SQL + schema introspection."""
    insp = inspect(engine)
    if table not in insp.get_table_names():
        return [], []

    cols_meta = insp.get_columns(table)
    cols = [c["name"] for c in cols_meta]
    if not cols:
        return [], []

    # Prefer deterministic ordering by id if present.
    order_clause = ""
    if "id" in cols:
        order_clause = ' ORDER BY "id" ASC'
    elif "ts" in cols:
        order_clause = ' ORDER BY "ts" ASC'

    limit_clause = ""
    if max_rows is not None:
        limit_clause = " LIMIT :_limit"

    quoted_cols = ", ".join([f'"{c}"' for c in cols])
    q = f'SELECT {quoted_cols} FROM "{table}"{order_clause}{limit_clause}'

    params = {"_limit": int(max_rows)} if max_rows is not None else {}
    with engine.begin() as conn:
        rows = conn.execute(sql_text(q), params).mappings().all()
    return cols, [dict(r) for r in rows]


def export_excel_dynamic(engine: Engine) -> Tuple[bytes, Dict[str, Any]]:
    """Export all tables to an Excel workbook. Returns (xlsx_bytes, report)."""
    wb = Workbook()
    wb.remove(wb.active)

    report: Dict[str, Any] = {
        "exported_at": datetime.utcnow().replace(tzinfo=timezone.utc).isoformat(),
        "tables": [],
        "errors": [],
    }

    tables = list_tables(engine)
    # Keep this first if present.
    preferred = [
        "movie",
        "assignment",
        "translation_task",
        "translation_submission",
        "vo_role_submission",
        "translator",
        "vo_team",
        "admin_user",
        "admin_telegram_user",
        "system_logs",
    ]
    ordered = [t for t in preferred if t in tables] + [t for t in tables if t not in preferred]

    max_logs = int(os.getenv("EXPORT_MAX_LOGS", "5000") or "5000")

    for t in ordered:
        try:
            cols, rows = fetch_table(engine, t, max_rows=(max_logs if t == "system_logs" else None))
            if not cols:
                continue

            # Excel sheet name max 31 chars
            sheet_name = (t[:31])
            # Ensure unique name if collision.
            base = sheet_name
            i = 2
            while sheet_name in wb.sheetnames:
                suffix = f"_{i}"
                sheet_name = (base[: (31 - len(suffix))] + suffix)
                i += 1

            ws = wb.create_sheet(sheet_name)
            # headers
            for c, h in enumerate(cols, start=1):
                ws.cell(row=1, column=c, value=h)
            # rows
            r_i = 2
            for row in rows:
                for c_i, col in enumerate(cols, start=1):
                    ws.cell(row=r_i, column=c_i, value=_safe_cell(row.get(col)))
                r_i += 1

            report["tables"].append({"table": t, "sheet": sheet_name, "rows": len(rows)})
        except Exception as e:
            report["errors"].append({"table": t, "error": str(e)})

    # Add Export_Info sheet
    ws_i = wb.create_sheet("Export_Info", 0)
    ws_i["A1"] = "exported_at"
    ws_i["B1"] = report["exported_at"]
    ws_i["A2"] = "tables_exported"
    ws_i["B2"] = len(report["tables"])
    ws_i["A3"] = "errors"
    ws_i["B3"] = len(report["errors"])
    if report["errors"]:
        ws_i["A5"] = "error_details"
        ws_i["A6"] = json.dumps(report["errors"], ensure_ascii=False, indent=2)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.read(), report


def backup_json_zip_dynamic(engine: Engine, app_version: str = "") -> Tuple[bytes, Dict[str, Any]]:
    """Export tables as per-table JSON files inside a ZIP.

    Returns (zip_bytes, report).
    """
    report: Dict[str, Any] = {
        "exported_at": datetime.utcnow().replace(tzinfo=timezone.utc).isoformat(),
        "app_version": app_version,
        "tables": [],
        "errors": [],
    }

    tables = list_tables(engine)
    max_logs = int(os.getenv("EXPORT_MAX_LOGS", "5000") or "5000")

    zbio = BytesIO()
    with zipfile.ZipFile(zbio, "w", compression=zipfile.ZIP_DEFLATED) as z:
        # per-table json
        for t in tables:
            try:
                cols, rows = fetch_table(engine, t, max_rows=(max_logs if t == "system_logs" else None))
                # normalize values
                norm_rows = []
                for r in rows:
                    nr = {k: _iso(v) for k, v in r.items()}
                    norm_rows.append(nr)
                payload = {"table": t, "columns": cols, "rows": norm_rows}
                z.writestr(f"tables/{t}.json", json.dumps(payload, ensure_ascii=False, indent=2))
                report["tables"].append({"table": t, "rows": len(rows)})
            except Exception as e:
                report["errors"].append({"table": t, "error": str(e)})

        z.writestr("meta.json", json.dumps(report, ensure_ascii=False, indent=2))

    zbio.seek(0)
    return zbio.read(), report
