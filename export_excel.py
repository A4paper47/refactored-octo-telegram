import os
from datetime import datetime
from typing import Optional, List, Tuple

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook

from models import Assignment, TranslationSubmission


# -----------------------------
# Helpers
# -----------------------------
def _safe_str(x) -> str:
    return "" if x is None else str(x)

def _dt_str(dt) -> str:
    if not dt:
        return ""
    try:
        return dt.strftime("%Y-%m-%d %H:%M")
    except Exception:
        return str(dt)

def _norm(s: str) -> str:
    return (s or "").strip().lower()

def _find_sheet(wb: Workbook, names: List[str]):
    for n in names:
        if n in wb.sheetnames:
            return wb[n]
    return None

def _clear_sheet(ws, keep_header_rows: int = 1):
    """
    Clear sheet content while keeping first N rows (headers).
    """
    max_row = ws.max_row
    if max_row <= keep_header_rows:
        return
    ws.delete_rows(keep_header_rows + 1, max_row - keep_header_rows)

def _write_table(ws, start_row: int, headers: List[str], rows: List[List]):
    # headers
    for c, h in enumerate(headers, start=1):
        ws.cell(row=start_row, column=c, value=h)

    r = start_row + 1
    for row in rows:
        for c, val in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=val)
        r += 1


# -----------------------------
# Export builders
# -----------------------------
def build_assignments_rows(project: Optional[str] = None) -> List[List]:
    q = Assignment.query
    if project:
        q = q.filter_by(project=project)

    items = q.order_by(Assignment.project.asc(), Assignment.created_at.asc()).all()

    out = []
    for a in items:
        out.append([
            a.project,
            a.vo,
            a.role,
            int(a.lines or 0),
            _dt_str(a.created_at),
        ])
    return out

def build_submissions_rows(movie: Optional[str] = None) -> List[List]:
    q = TranslationSubmission.query
    if movie:
        q = q.filter_by(movie=movie)

    items = q.order_by(TranslationSubmission.submitted_at.asc()).all()

    out = []
    for s in items:
        out.append([
            s.movie,
            s.status,
            _safe_str(getattr(s, "submitter_name", "") or s.submitter_username or s.submitter_id),
            _safe_str(s.submitter_id),
            s.content_type,
            _safe_str(getattr(s, "category", "")),
            _safe_str(s.file_name),
            _safe_str(getattr(s, "mime_type", "")),
            _safe_str(getattr(s, "file_size", "")),
            _safe_str(getattr(s, "submitted_local", "")) or _dt_str(s.submitted_at),
            _safe_str(getattr(s, "note", "")),
        ])
    return out


# -----------------------------
# Smart template fill (best effort)
# -----------------------------
COMMON_HEADERS = {
    "movie": ["movie", "title", "film", "project"],
    "role": ["role", "character", "cast"],
    "lines": ["lines", "line", "total lines", "jumlah line"],
    "vo": ["vo", "voice", "talent", "actor", "speaker", "person"],
    "translator": ["translator", "tl", "trans"],
    "deadline": ["deadline", "due", "due date", "dateline"],
    "status": ["status", "stage"],
}

def _scan_header_row(ws, row: int, max_cols: int = 80):
    """
    Return dict: header_key -> column_index if matched by header text.
    """
    found = {}
    for col in range(1, max_cols + 1):
        v = ws.cell(row=row, column=col).value
        if not v:
            continue
        t = _norm(str(v))
        for key, alts in COMMON_HEADERS.items():
            if key in found:
                continue
            if any(t == _norm(a) for a in alts):
                found[key] = col
    return found

def _find_best_header(ws, max_rows: int = 40):
    """
    Try to find a header row within first N rows.
    Return (header_row_index, mapping dict) or (None, {}).
    """
    best = (None, {})
    for r in range(1, max_rows + 1):
        mapping = _scan_header_row(ws, r)
        # we consider "good" if it has at least movie + lines (or movie + role)
        score = len(mapping)
        if score > len(best[1]):
            best = (r, mapping)
    return best


def export_with_template(
    template_path: str,
    out_path: str,
    project_filter: Optional[str] = None,
    movie_filter: Optional[str] = None,
):
    """
    Loads Excel template and writes exports.

    - If template has recognizable table headers, fills them.
    - Always creates fallback sheets Export_Assignments and Export_Submissions.
    """
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Template not found: {template_path}")

    wb = load_workbook(template_path)

    # --- Always create/refresh fallback sheets ---
    if "Export_Assignments" in wb.sheetnames:
        ws_a = wb["Export_Assignments"]
        wb.remove(ws_a)
    ws_a = wb.create_sheet("Export_Assignments", 0)

    if "Export_Submissions" in wb.sheetnames:
        ws_s = wb["Export_Submissions"]
        wb.remove(ws_s)
    ws_s = wb.create_sheet("Export_Submissions", 1)

    # data
    assignment_rows = build_assignments_rows(project_filter)
    submission_rows = build_submissions_rows(movie_filter)

    _write_table(
        ws_a,
        start_row=1,
        headers=["Movie/Project", "VO", "Role", "Lines", "Created At"],
        rows=assignment_rows
    )

    _write_table(
        ws_s,
        start_row=1,
        headers=[
            "Movie", "Status", "Submitter", "Submitter ID",
            "Content Type", "Category", "File Name", "MIME", "File Size",
            "Submitted", "Note"
        ],
        rows=submission_rows
    )

    # --- Best-effort fill into existing tracker sheet if possible ---
    ws_tracker = _find_sheet(wb, ["Films Tracker", "Film Tracker", "Tracker", "Films", "Sheet1"])
    if ws_tracker:
        header_row, mapping = _find_best_header(ws_tracker)
        # only attempt if we at least find movie column
        if header_row and "movie" in mapping:
            # clear rows after header
            _clear_sheet(ws_tracker, keep_header_rows=header_row)

            # write assignments into tracker (one row per assignment)
            write_row = header_row + 1
            for a in Assignment.query.order_by(Assignment.created_at.asc()).all():
                # movie
                ws_tracker.cell(row=write_row, column=mapping["movie"], value=a.project)

                # optional columns
                if "vo" in mapping:
                    ws_tracker.cell(row=write_row, column=mapping["vo"], value=a.vo)
                if "role" in mapping:
                    ws_tracker.cell(row=write_row, column=mapping["role"], value=a.role)
                if "lines" in mapping:
                    ws_tracker.cell(row=write_row, column=mapping["lines"], value=int(a.lines or 0))
                if "status" in mapping:
                    ws_tracker.cell(row=write_row, column=mapping["status"], value="ASSIGNED")

                write_row += 1

    # meta sheet (nice to have)
    if "Export_Info" in wb.sheetnames:
        ws_i = wb["Export_Info"]
        wb.remove(ws_i)
    ws_i = wb.create_sheet("Export_Info", 2)
    ws_i["A1"] = "Exported At"
    ws_i["B1"] = datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")
    ws_i["A2"] = "Project Filter"
    ws_i["B2"] = project_filter or ""
    ws_i["A3"] = "Movie Filter"
    ws_i["B3"] = movie_filter or ""

    wb.save(out_path)
    


# -----------------------------
# Simple full export (no template)
# -----------------------------
def export_simple(
    out_path: str,
    project_filter: Optional[str] = None,
    movie_filter: Optional[str] = None,
):
    """Create a standalone Excel export (safe for backups).

    Sheets:
      - Movies
      - Assignments
      - VO_Submissions
      - Queue (Translation_Submissions)
      - VO_Team
      - Translators

    This works on both SQLite and Postgres and does not depend on any template file.
    """
    from openpyxl import Workbook
    from models import Movie, VORoleSubmission, VOTeam, Translator

    wb = Workbook()
    # remove default sheet
    wb.remove(wb.active)

    # Movies
    ws_m = wb.create_sheet("Movies", 0)
    movies = Movie.query.order_by(Movie.created_at.asc()).all()
    rows_m = []
    for m in movies:
        rows_m.append([
            _safe_str(m.code),
            _safe_str(m.title),
            _safe_str(m.year),
            _safe_str(m.lang),
            _safe_str(m.status),
            _dt_str(m.received_at),
            _dt_str(m.submitted_at),
            _dt_str(m.completed_at),
            _safe_str(m.translator_assigned),
            _safe_str(m.vo_group_chat_id),
            _dt_str(m.created_at),
            _dt_str(m.updated_at),
        ])
    _write_table(
        ws_m, 1,
        headers=[
            "Code", "Title", "Year", "Lang", "Status",
            "Received At", "Submitted At", "Completed At",
            "Translator Assigned", "VO Group Chat ID",
            "Created At", "Updated At",
        ],
        rows=rows_m,
    )

    # Assignments
    ws_a = wb.create_sheet("Assignments", 1)
    assignment_rows = build_assignments_rows(project_filter)
    _write_table(
        ws_a, 1,
        headers=["Movie/Project", "VO", "Role", "Lines", "Created At"],
        rows=assignment_rows,
    )

    # VO submissions
    ws_vo = wb.create_sheet("VO_Submissions", 2)
    q_vo = VORoleSubmission.query
    if movie_filter:
        q_vo = q_vo.filter_by(movie=movie_filter)
    subs_vo = q_vo.order_by(VORoleSubmission.submitted_at.asc()).all()
    rows_vo = []
    for s in subs_vo:
        rows_vo.append([
            _safe_str(s.movie),
            _safe_str(s.role),
            _safe_str(s.vo),
            int(s.lines or 0),
            _dt_str(s.submitted_at),
            _safe_str(getattr(s, "file_name", "")),
            _safe_str(getattr(s, "media_type", "")),
            _safe_str(getattr(s, "tg_chat_id", "")),
            _safe_str(getattr(s, "tg_message_id", "")),
        ])
    _write_table(
        ws_vo, 1,
        headers=[
            "Movie", "Role", "VO", "Lines", "Submitted At",
            "File Name", "Media Type", "TG Chat ID", "TG Message ID"
        ],
        rows=rows_vo,
    )

    # Queue (translation submissions)
    ws_q = wb.create_sheet("Queue", 3)
    submission_rows = build_submissions_rows(movie_filter)
    _write_table(
        ws_q, 1,
        headers=[
            "Movie", "Status", "Submitter", "Submitter ID",
            "Content Type", "Category", "File Name", "MIME", "File Size",
            "Submitted", "Note"
        ],
        rows=submission_rows,
    )

    # VO Team
    ws_team = wb.create_sheet("VO_Team", 4)
    team = VOTeam.query.order_by(VOTeam.gender.asc(), VOTeam.name.asc()).all()
    rows_team = []
    for v in team:
        rows_team.append([
            _safe_str(v.name),
            _safe_str(v.gender),
            _safe_str(v.level),
            _safe_str(v.speed),
            int(bool(getattr(v, "urgent_ok", False))),
            int(bool(getattr(v, "active", True))),
        ])
    _write_table(
        ws_team, 1,
        headers=["Name", "Gender", "Level", "Speed", "Urgent OK", "Active"],
        rows=rows_team,
    )

    # Translators
    ws_tr = wb.create_sheet("Translators", 5)
    trs = Translator.query.order_by(Translator.active.desc(), Translator.name.asc()).all()
    rows_tr = []
    for t in trs:
        rows_tr.append([
            _safe_str(t.name),
            _safe_str(t.tg_username),
            _safe_str(t.tg_user_id),
            int(bool(getattr(t, "active", True))),
            _safe_str(getattr(t, "languages", "")),
            _safe_str(getattr(t, "note", "")),
            _dt_str(getattr(t, "last_seen_at", None)),
            _dt_str(getattr(t, "created_at", None)),
            _dt_str(getattr(t, "updated_at", None)),
        ])
    _write_table(
        ws_tr, 1,
        headers=["Name", "TG Username", "TG User ID", "Active", "Languages", "Note", "Last Seen", "Created At", "Updated At"],
        rows=rows_tr,
    )

    wb.save(out_path)
